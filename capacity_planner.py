"""
Capacity Planning Tool
Reads work requests from Excel, calculates weekly/monthly capacity for the team,
and outputs professional Gantt, capacity, and roadmap charts as PNGs.

Features:
  - Priority-based ordering (P1-P4) for workstreams and tasks
  - Planned vs actual tracking with estimation drift
  - Fractional day support (0.5, 1.5, etc.)
  - Blocked/On Hold tracking with duration and reason
  - Semi-automated schedule suggestions
  - Enhanced Excel template with dropdowns and conditional formatting
"""

import argparse
import difflib
import io
import math
import os
import re
import sys
from calendar import monthrange
from datetime import datetime, timedelta

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import matplotlib.patheffects as pe
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule


# ── Constants ────────────────────────────────────────────────────────────────

_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INPUT = os.path.join(_DIR, "capacity_data.xlsx")
DEFAULT_OUTPUT = os.path.join(_DIR, "output", "capacity_gantt.png")
DEFAULT_MONTHLY_OUTPUT = os.path.join(_DIR, "output", "capacity_monthly.png")
DEFAULT_ROADMAP_OUTPUT = os.path.join(_DIR, "output", "roadmap.png")
DEFAULT_WEEKLY_OUTPUT = os.path.join(_DIR, "output", "capacity_weekly.png")

# These defaults are ONLY used when generating the Excel template via --template.
# At runtime, all workstream names, colors, and priorities are loaded from the Excel file.
WORKSTREAM_COLORS = {
    "Strategic Initiative A": "#2196F3",
    "Strategic Initiative B": "#F44336",
    "Operational Process A": "#4CAF50",
    "Operational Process B": "#FF9800",
    "Platform Migration Alpha": "#9C27B0",
    "Platform Migration Beta": "#607D8B",
    "Continuous Delivery": "#00BCD4",
    "Team Development": "#FF5722",
    "Infrastructure Modernisation": "#795548",
}

WORKSTREAM_PRIORITIES = {
    "Strategic Initiative A": "P1",
    "Strategic Initiative B": "P1",
    "Operational Process A": "P2",
    "Operational Process B": "P2",
    "Platform Migration Alpha": "P3",
    "Platform Migration Beta": "P4",
    "Continuous Delivery": "P1",
    "Team Development": "P2",
    "Infrastructure Modernisation": "P3",
}

STATUS_VALUES = ["Planned", "In Progress", "Complete", "On Hold"]
PRIORITY_VALUES = ["P1", "P2", "P3", "P4"]
LEAVE_TYPES = ["Annual Leave", "Sick", "Training", "Conference", "Other"]
CONFIDENCE_VALUES = ["High", "Medium", "Low"]
CONFIDENCE_COLORS = {"Low": "#E53935", "Medium": "#FF8F00", "High": "#43A047"}

PERSON_HATCHES = {
    0: "",       # First person: solid fill
    1: "//",     # Second person: diagonal hatch
    2: "\\\\",   # Third: back-diagonal
    3: "xx",     # Fourth: cross-hatch
}

STATUS_SYMBOLS = {
    "In Progress": "\u2022",   # bullet
    "Planned": "\u2013",       # en dash
    "Complete": "\u2714",      # heavy check mark
    "On Hold": "\u2016",       # double vertical line
}

PRIORITY_STYLES = {
    "P1": {"linewidth": 2.2, "label_weight": "bold",   "alpha": 1.0,  "label_size_bump": 0.5},
    "P2": {"linewidth": 1.5, "label_weight": "medium", "alpha": 0.95, "label_size_bump": 0},
    "P3": {"linewidth": 1.0, "label_weight": "normal", "alpha": 0.80, "label_size_bump": 0},
    "P4": {"linewidth": 0.8, "label_weight": "normal", "alpha": 0.65, "label_size_bump": -0.5},
}

STYLE = {
    "font_family": "Segoe UI",
    "title_size": 18,
    "subtitle_size": 13,
    "label_size": 9.5,
    "tick_size": 8.5,
    "small_size": 7.5,
    "bg_color": "#FAFAFA",
    "panel_bg": "#FFFFFF",
    "text_primary": "#1A1A2E",
    "text_secondary": "#555555",
    "text_muted": "#999999",
    "grid_color": "#E0E0E0",
    "today_color": "#D32F2F",
    "over_capacity_color": "#E53935",
    "under_capacity_colors": ["#43A047", "#1E88E5", "#8E24AA", "#FB8C00"],
    "capacity_line_color": "#1A1A2E",
    "row_shade_even": "#F5F5F5",
    "row_shade_odd": "#FFFFFF",
    "header_bg_alpha": 0.08,
    "bar_height": 0.6,
    "dpi": 180,
    "fig_width": 20,
    "drift_increase_color": "#FF8F00",
    "drift_decrease_color": "#43A047",
    "early_color": "#43A047",
    "late_color": "#E53935",
    "on_hold_color": "#B0BEC5",        # Blue-grey — neutral "parked" look
    "on_hold_edge_color": "#37474F",   # Dark blue-grey for strong contrast
    "planned_alpha_factor": 0.7,       # Planned bars = priority alpha × this
    "leave_color": "#FFF9C4",          # Pale yellow for leave markers
    "leave_edge_color": "#F9A825",     # Amber for leave text/markers
    "holiday_color": "#E1BEE7",        # Pale purple for public holiday shading
    "holiday_edge_color": "#7B1FA2",   # Deep purple for holiday markers
}


# ── Style Helpers ────────────────────────────────────────────────────────────

def apply_style():
    """Configure matplotlib rcParams for consistent styling."""
    plt.rcParams.update({
        "font.family": STYLE["font_family"],
        "font.size": STYLE["label_size"],
        "axes.facecolor": STYLE["panel_bg"],
        "figure.facecolor": STYLE["bg_color"],
        "axes.edgecolor": STYLE["grid_color"],
        "axes.linewidth": 0.8,
        "axes.grid": False,
        "grid.color": STYLE["grid_color"],
        "grid.linewidth": 0.5,
        "grid.alpha": 0.3,
        "xtick.color": STYLE["text_secondary"],
        "ytick.color": STYLE["text_secondary"],
        "text.color": STYLE["text_primary"],
    })


def style_axes(ax, title="", ylabel="", show_grid_x=False, show_grid_y=False):
    """Apply consistent axis styling to any subplot."""
    if title:
        ax.set_title(title, fontsize=STYLE["subtitle_size"], fontweight="bold",
                     color=STYLE["text_primary"], pad=12, loc="left")
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=STYLE["label_size"], color=STYLE["text_secondary"])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_linewidth(0.6)
    ax.spines["bottom"].set_linewidth(0.6)
    if show_grid_x:
        ax.grid(axis="x", alpha=0.15, linewidth=0.5, color=STYLE["grid_color"])
    if show_grid_y:
        ax.grid(axis="y", alpha=0.15, linewidth=0.5, color=STYLE["grid_color"])
    ax.set_axisbelow(True)


def add_header_footer(fig, title, subtitle="", data_mtime=None):
    """Add a professional title block and generation timestamp footer."""
    fig.suptitle(title, fontsize=STYLE["title_size"], fontweight="bold",
                 color=STYLE["text_primary"], y=0.98, x=0.04, ha="left")
    if subtitle:
        fig.text(0.04, 0.948, subtitle, fontsize=STYLE["small_size"] + 1,
                 color=STYLE["text_muted"], ha="left")
    fig.text(0.98, 0.008, f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}",
             ha="right", fontsize=STYLE["small_size"], color=STYLE["text_muted"])
    footer_left = "Capacity Planning Tool"
    mtime_str = data_mtime or STYLE.get("_data_mtime")
    if mtime_str:
        footer_left += f"  \u00b7  Data updated {mtime_str}"
    fig.text(0.04, 0.008, footer_left,
             ha="left", fontsize=STYLE["small_size"], color=STYLE["text_muted"])


def draw_today_line(ax, date_min, date_max, y_top):
    """Draw a styled 'Today' marker if today falls within the date range."""
    today = datetime.now()
    today_num = mdates.date2num(today)
    min_num = mdates.date2num(date_min)
    max_num = mdates.date2num(date_max)
    if min_num <= today_num <= max_num:
        ax.axvspan(today_num - 0.3, today_num + 0.3,
                   color=STYLE["today_color"], alpha=0.06, zorder=1)
        ax.axvline(today_num, color=STYLE["today_color"], linewidth=2,
                   linestyle="-", alpha=0.7, zorder=10)
        ax.text(today_num + 0.5, y_top, "Today", fontsize=STYLE["small_size"] + 0.5,
                color=STYLE["today_color"], fontweight="bold", va="bottom",
                ha="left", style="italic")


def draw_rounded_bar(ax, x, y, width, height, color, alpha=1.0,
                     edgecolor=None, linewidth=1.2, hatch="", zorder=3,
                     linestyle="-"):
    """Draw a horizontal bar with rounded corners using FancyBboxPatch."""
    if width <= 0:
        return None
    rounding = min(0.12, height * 0.3, width * 0.05)
    fancy = FancyBboxPatch(
        (x, y - height / 2), width, height,
        boxstyle=f"round,pad=0,rounding_size={rounding}",
        facecolor=color, alpha=alpha,
        edgecolor=edgecolor or color, linewidth=linewidth,
        hatch=hatch, zorder=zorder, linestyle=linestyle,
    )
    ax.add_patch(fancy)
    return fancy


def priority_sort_key(priority_str):
    """Convert 'P1'->1, 'P2'->2, etc. for sorting."""
    if priority_str and len(priority_str) == 2 and priority_str[0] == "P":
        try:
            return int(priority_str[1])
        except ValueError:
            pass
    return 9


def norm_date(d):
    """Normalise to midnight datetime for safe set membership checks."""
    if not isinstance(d, (datetime, pd.Timestamp)):
        raise TypeError(f"norm_date expected datetime, got {type(d).__name__}: {d!r}")
    if isinstance(d, pd.Timestamp):
        d = d.to_pydatetime()
    return datetime(d.year, d.month, d.day)


def clean_str(val):
    """Return stripped string or empty string for NaN/None."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    return str(val).strip()


class _TeeWriter:
    """Write to two streams simultaneously (for summary.txt capture)."""
    def __init__(self, a, b):
        self.a, self.b = a, b
    def write(self, data):
        self.a.write(data)
        self.b.write(data)
    def flush(self):
        self.a.flush()
        self.b.flush()


def parse_date(val, context=""):
    """Parse date from Excel cell — handles datetime, Timestamp, and string."""
    if pd.isna(val):
        raise ValueError(f"Date is blank{f' ({context})' if context else ''}")
    if isinstance(val, datetime):
        return norm_date(val)
    if isinstance(val, pd.Timestamp):
        return norm_date(val.to_pydatetime())
    if isinstance(val, str):
        val = val.strip()
        if not val:
            raise ValueError(f"Date is blank{f' ({context})' if context else ''}")
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return norm_date(datetime.strptime(val, fmt))
            except ValueError:
                pass
        ctx = f" ({context})" if context else ""
        raise ValueError(f"Cannot parse date{ctx}: {val!r}. Expected YYYY-MM-DD or DD/MM/YYYY")
    raise ValueError(f"Cannot parse date{f' ({context})' if context else ''}: {val!r}")


def is_working_day(date, public_holidays=None, person_leave=None):
    """Check if a date is a working day (not weekend, not public holiday, not on leave)."""
    if date.weekday() >= 5:
        return False
    d_normalised = norm_date(date)
    if public_holidays and d_normalised in public_holidays:
        return False
    if person_leave and d_normalised in person_leave:
        return False
    return True


def count_working_days(start, end, public_holidays=None, person_leave=None):
    """Count working days between start and end (inclusive)."""
    d, end_d = norm_date(start), norm_date(end)
    count = 0
    while d <= end_d:
        if is_working_day(d, public_holidays, person_leave):
            count += 1
        d += timedelta(days=1)
    return count


def working_days_in_month(year, month, public_holidays=None):
    """Count working days in a month (weekdays minus public holidays)."""
    _, num_days = monthrange(year, month)
    count = 0
    for d in range(1, num_days + 1):
        dt = datetime(year, month, d)
        if dt.weekday() < 5 and (not public_holidays or dt not in public_holidays):
            count += 1
    return count


# ── Template Generation ─────────────────────────────────────────────────────

def generate_template(output_path):
    """Create an Excel template with 5 sheets (Team, Workstreams, Tasks, Public Holidays, Leave),
    example data, dropdowns, and conditional formatting."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E3B4E", end_color="2E3B4E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data_rows(ws, start_row=2):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    team_names = ["Team Lead", "Analyst"]
    ws_names_list = list(WORKSTREAM_COLORS.keys())

    # ── Sheet 1: Team ──
    ws_team = wb.active
    ws_team.title = "Team"
    ws_team.append(["Name", "Role", "Days Per Week"])
    ws_team.append(["Team Lead", "Lead", 5])
    ws_team.append(["Analyst", "Analyst", 5])
    ws_team.column_dimensions["A"].width = 20
    ws_team.column_dimensions["B"].width = 15
    ws_team.column_dimensions["C"].width = 16
    style_header(ws_team)
    style_data_rows(ws_team)
    ws_team.freeze_panes = "A2"

    # ── Sheet 2: Workstreams ──
    ws_workstreams = wb.create_sheet("Workstreams")
    ws_workstreams.append(["Workstream", "Color", "Priority"])
    for ws_name, color in WORKSTREAM_COLORS.items():
        priority = WORKSTREAM_PRIORITIES.get(ws_name, "P2")
        ws_workstreams.append([ws_name, color, priority])
    ws_workstreams.column_dimensions["A"].width = 45
    ws_workstreams.column_dimensions["B"].width = 12
    ws_workstreams.column_dimensions["C"].width = 12
    style_header(ws_workstreams)
    style_data_rows(ws_workstreams)
    ws_workstreams.freeze_panes = "A2"

    # Color preview fills
    for row_idx in range(2, ws_workstreams.max_row + 1):
        color_cell = ws_workstreams.cell(row=row_idx, column=2)
        hex_color = color_cell.value.lstrip("#") if color_cell.value else "FFFFFF"
        color_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Priority dropdown on Workstreams sheet
    dv_ws_priority = DataValidation(type="list", formula1='"P1,P2,P3,P4"', allow_blank=False)
    dv_ws_priority.error = "Please select P1, P2, P3, or P4"
    dv_ws_priority.errorTitle = "Invalid Priority"
    ws_workstreams.add_data_validation(dv_ws_priority)
    for row_idx in range(2, ws_workstreams.max_row + 1):
        dv_ws_priority.add(ws_workstreams.cell(row=row_idx, column=3))

    # Priority conditional formatting on Workstreams sheet
    ws_workstreams.conditional_formatting.add(
        f"C2:C{ws_workstreams.max_row}",
        CellIsRule(operator="equal", formula=['"P1"'],
                   font=Font(bold=True, color="C62828"), fill=PatternFill(bgColor="FFCDD2")))
    ws_workstreams.conditional_formatting.add(
        f"C2:C{ws_workstreams.max_row}",
        CellIsRule(operator="equal", formula=['"P2"'],
                   font=Font(bold=True, color="E65100"), fill=PatternFill(bgColor="FFE0B2")))
    ws_workstreams.conditional_formatting.add(
        f"C2:C{ws_workstreams.max_row}",
        CellIsRule(operator="equal", formula=['"P4"'],
                   font=Font(color="9E9E9E"), fill=PatternFill(bgColor="F5F5F5")))

    # ── Sheet 3: Tasks ──
    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append([
        "Task", "Workstream", "Assigned To", "Start Date",
        "Original Days", "Total Days", "Priority", "Status",
        "Actual End", "Blocked By", "Deadline", "Confidence", "Notes",
    ])

    # Example tasks: Priority blank on some to demonstrate inheritance; Deadline/Confidence on 1-2
    example_tasks = [
        ["Requirements Gathering", "Strategic Initiative A", "Team Lead",
         "2026-02-16", 10, 10, "P1", "In Progress", "", "", "", "", "Discovery phase with stakeholders"],
        ["Data Analysis Sprint", "Continuous Delivery", "Analyst",
         "2026-02-16", 8, 8, "", "In Progress", "", "", "", "", "Monthly reporting cycle"],
        ["Market Research Report", "Strategic Initiative B", "Team Lead",
         "2026-03-02", 15, 15, "", "Planned", "", "", "2026-03-20", "Medium", "Competitor analysis — board deadline"],
        ["Weekly Team Sync", "Team Development", "Team Lead",
         "2026-02-16", 2, 2.5, "", "In Progress", "", "", "", "", "Recurring - 0.5 days/week"],
        ["Process Documentation", "Operational Process A", "Team Lead",
         "2026-03-16", 5, 5, "", "Planned", "", "", "", "", "Document current workflows"],
        ["Training Materials Update", "Operational Process B", "Analyst",
         "2026-02-23", 3, 3, "", "Planned", "", "", "", "", "Q2 onboarding refresh"],
        ["Platform Evaluation", "Platform Migration Alpha", "Team Lead",
         "2026-04-01", 20, 20, "P1", "Planned", "", "", "2026-05-01", "Low", "Vendor assessment — new tech, uncertain scope"],
        ["Dashboard Migration Phase 1", "Platform Migration Beta", "Analyst",
         "2026-03-16", 15, 15, "", "Planned", "", "", "", "", "Migrate top-used dashboards"],
        ["Data Pipeline Review", "Infrastructure Modernisation", "Team Lead",
         "2026-03-09", 10, 12, "", "Planned", "", "", "", "", "Audit existing ETL pipelines"],
        ["Proof of Concept Build", "Platform Migration Beta", "Analyst",
         "2026-05-01", 12, 12, "", "Planned", "", "", "", "High", "Initial technical spike"],
    ]
    for task in example_tasks:
        ws_tasks.append(task)

    ws_tasks.column_dimensions["A"].width = 35
    ws_tasks.column_dimensions["B"].width = 42
    ws_tasks.column_dimensions["C"].width = 14
    ws_tasks.column_dimensions["D"].width = 14
    ws_tasks.column_dimensions["E"].width = 14
    ws_tasks.column_dimensions["F"].width = 12
    ws_tasks.column_dimensions["G"].width = 10
    ws_tasks.column_dimensions["H"].width = 13
    ws_tasks.column_dimensions["I"].width = 14
    ws_tasks.column_dimensions["J"].width = 25
    ws_tasks.column_dimensions["K"].width = 14   # Deadline
    ws_tasks.column_dimensions["L"].width = 13   # Confidence
    ws_tasks.column_dimensions["M"].width = 35   # Notes
    style_header(ws_tasks)
    style_data_rows(ws_tasks)
    ws_tasks.freeze_panes = "A2"

    # Center date columns (D=Start Date, I=Actual End, K=Deadline)
    for row_idx in range(2, ws_tasks.max_row + 1):
        ws_tasks.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws_tasks.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center", vertical="center")
        ws_tasks.cell(row=row_idx, column=11).alignment = Alignment(horizontal="center", vertical="center")

    # ── Data Validations on Tasks sheet ──
    max_task_row = 100  # allow room for future rows

    # Status dropdown
    dv_status = DataValidation(type="list", formula1='"Planned,In Progress,Complete,On Hold"', allow_blank=False)
    dv_status.error = "Please select a valid status"
    dv_status.errorTitle = "Invalid Status"
    ws_tasks.add_data_validation(dv_status)
    dv_status.add(f"H2:H{max_task_row}")

    # Priority dropdown (allow_blank=True for workstream inheritance)
    dv_task_priority = DataValidation(type="list", formula1='"P1,P2,P3,P4"', allow_blank=True)
    dv_task_priority.error = "Please select P1, P2, P3, P4 — or leave blank to inherit from workstream"
    dv_task_priority.errorTitle = "Invalid Priority"
    ws_tasks.add_data_validation(dv_task_priority)
    dv_task_priority.add(f"G2:G{max_task_row}")

    # Workstream dropdown (range-based to avoid 255-char limit with long names)
    dv_workstream = DataValidation(type="list", formula1="=Workstreams!$A$2:$A$100", allow_blank=False)
    dv_workstream.error = "Please select a valid workstream"
    dv_workstream.errorTitle = "Invalid Workstream"
    ws_tasks.add_data_validation(dv_workstream)
    dv_workstream.add(f"B2:B{max_task_row}")

    # Assigned To dropdown (range-based to avoid 255-char limit)
    dv_assigned = DataValidation(type="list", formula1="=Team!$A$2:$A$50", allow_blank=False)
    dv_assigned.error = "Please select a team member"
    dv_assigned.errorTitle = "Invalid Team Member"
    ws_tasks.add_data_validation(dv_assigned)
    dv_assigned.add(f"C2:C{max_task_row}")

    # Confidence dropdown (allow_blank=True — optional)
    dv_confidence = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    dv_confidence.error = "Please select High, Medium, Low — or leave blank"
    dv_confidence.errorTitle = "Invalid Confidence"
    ws_tasks.add_data_validation(dv_confidence)
    dv_confidence.add(f"L2:L{max_task_row}")

    # ── Conditional Formatting on Tasks sheet ──
    # Status colours
    status_range = f"H2:H{max_task_row}"
    ws_tasks.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Complete"'],
                   font=Font(color="1B5E20"), fill=PatternFill(bgColor="C8E6C9")))
    ws_tasks.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"In Progress"'],
                   font=Font(color="E65100"), fill=PatternFill(bgColor="FFE0B2")))
    ws_tasks.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"On Hold"'],
                   font=Font(color="B71C1C"), fill=PatternFill(bgColor="FFCDD2")))
    ws_tasks.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Planned"'],
                   font=Font(color="757575"), fill=PatternFill(bgColor="F5F5F5")))

    # Priority colours on Tasks
    priority_range = f"G2:G{max_task_row}"
    ws_tasks.conditional_formatting.add(
        priority_range,
        CellIsRule(operator="equal", formula=['"P1"'],
                   font=Font(bold=True, color="C62828"), fill=PatternFill(bgColor="FFCDD2")))
    ws_tasks.conditional_formatting.add(
        priority_range,
        CellIsRule(operator="equal", formula=['"P2"'],
                   font=Font(bold=True, color="E65100"), fill=PatternFill(bgColor="FFE0B2")))
    ws_tasks.conditional_formatting.add(
        priority_range,
        CellIsRule(operator="equal", formula=['"P4"'],
                   font=Font(color="9E9E9E"), fill=PatternFill(bgColor="F5F5F5")))

    # Confidence colours
    conf_range = f"L2:L{max_task_row}"
    ws_tasks.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"Low"'],
                   font=Font(bold=True, color="B71C1C"), fill=PatternFill(bgColor="FFCDD2")))
    ws_tasks.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"Medium"'],
                   font=Font(color="E65100"), fill=PatternFill(bgColor="FFE0B2")))
    ws_tasks.conditional_formatting.add(
        conf_range,
        CellIsRule(operator="equal", formula=['"High"'],
                   font=Font(color="1B5E20"), fill=PatternFill(bgColor="C8E6C9")))

    # Scope increase highlight: Total Days > Original Days (amber fill)
    # Column E = Original Days, Column F = Total Days
    ws_tasks.conditional_formatting.add(
        f"F2:F{max_task_row}",
        FormulaRule(formula=["AND(F2>E2,E2>0)"],
                    fill=PatternFill(bgColor="FFE0B2"),
                    font=Font(bold=True, color="E65100")))

    # ── Sheet 4: Public Holidays ──
    ws_holidays = wb.create_sheet("Public Holidays")
    ws_holidays.append(["Date", "Name"])
    example_holidays = [
        ["2026-01-01", "New Year's Day"],
        ["2026-04-03", "Good Friday"],
        ["2026-04-06", "Easter Monday"],
        ["2026-05-04", "Early May Bank Holiday"],
        ["2026-05-25", "Spring Bank Holiday"],
        ["2026-08-31", "Summer Bank Holiday"],
        ["2026-12-25", "Christmas Day"],
        ["2026-12-28", "Boxing Day (substitute)"],
    ]
    for hol in example_holidays:
        ws_holidays.append(hol)
    ws_holidays.column_dimensions["A"].width = 16
    ws_holidays.column_dimensions["B"].width = 30
    style_header(ws_holidays)
    style_data_rows(ws_holidays)
    ws_holidays.freeze_panes = "A2"

    # Center date column
    for row_idx in range(2, ws_holidays.max_row + 1):
        ws_holidays.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 5: Leave ──
    ws_leave = wb.create_sheet("Leave")
    ws_leave.append(["Person", "Start Date", "End Date", "Type", "Notes"])
    example_leave = [
        ["Team Lead", "2026-04-06", "2026-04-10", "Annual Leave", "Easter week"],
        ["Analyst", "2026-03-23", "2026-03-25", "Training", "Platform training course"],
        ["Team Lead", "2026-06-15", "2026-06-19", "Annual Leave", "Summer break"],
    ]
    for lv in example_leave:
        ws_leave.append(lv)
    ws_leave.column_dimensions["A"].width = 16
    ws_leave.column_dimensions["B"].width = 14
    ws_leave.column_dimensions["C"].width = 14
    ws_leave.column_dimensions["D"].width = 16
    ws_leave.column_dimensions["E"].width = 30
    style_header(ws_leave)
    style_data_rows(ws_leave)
    ws_leave.freeze_panes = "A2"

    # Center date columns
    for row_idx in range(2, ws_leave.max_row + 1):
        ws_leave.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws_leave.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")

    max_leave_row = 100

    # Person dropdown on Leave sheet (range-based)
    dv_leave_person = DataValidation(type="list", formula1="=Team!$A$2:$A$50", allow_blank=False)
    dv_leave_person.error = "Please select a team member"
    dv_leave_person.errorTitle = "Invalid Person"
    ws_leave.add_data_validation(dv_leave_person)
    dv_leave_person.add(f"A2:A{max_leave_row}")

    # Type dropdown on Leave sheet
    leave_type_str = ",".join(LEAVE_TYPES)
    dv_leave_type = DataValidation(type="list", formula1=f'"{leave_type_str}"', allow_blank=False)
    dv_leave_type.error = "Please select a valid leave type"
    dv_leave_type.errorTitle = "Invalid Leave Type"
    ws_leave.add_data_validation(dv_leave_type)
    dv_leave_type.add(f"D2:D{max_leave_row}")

    # Conditional formatting on Leave Type
    type_range = f"D2:D{max_leave_row}"
    ws_leave.conditional_formatting.add(
        type_range,
        CellIsRule(operator="equal", formula=['"Annual Leave"'],
                   font=Font(color="1B5E20"), fill=PatternFill(bgColor="C8E6C9")))
    ws_leave.conditional_formatting.add(
        type_range,
        CellIsRule(operator="equal", formula=['"Sick"'],
                   font=Font(color="B71C1C"), fill=PatternFill(bgColor="FFCDD2")))
    ws_leave.conditional_formatting.add(
        type_range,
        CellIsRule(operator="equal", formula=['"Training"'],
                   font=Font(color="E65100"), fill=PatternFill(bgColor="FFE0B2")))

    wb.save(output_path)
    print(f"Template created: {output_path}")
    print("  - Sheet 'Team': define your team members and available days")
    print("  - Sheet 'Workstreams': workstreams with display colors and priorities")
    print("  - Sheet 'Tasks': tasks with estimates, tracking, optional deadlines and confidence")
    print("  - Sheet 'Public Holidays': dates that affect all team members")
    print("  - Sheet 'Leave': per-person leave with dates and type")
    print("  - Dropdowns: Status, Priority, Workstream, Assigned To, Confidence, Leave Type")
    print("  - Conditional formatting: Status, Priority, Confidence, Scope drift, Leave types")
    print("  - Smart defaults: Priority inherits from workstream if blank; Original Days = Total Days if blank")
    print(f"\nEdit the file, then run again without --template to generate the charts.")


# ── Data Loading ─────────────────────────────────────────────────────────────

def load_team(filepath):
    """Load team members from the 'Team' sheet."""
    try:
        df = pd.read_excel(filepath, sheet_name="Team")
    except Exception as e:
        print(f"  WARNING: Could not read Team sheet: {e}")
        return {}
    if df.empty:
        return {}
    df.columns = df.columns.str.strip()
    required = {"Name", "Days Per Week"}
    missing = required - set(df.columns)
    if missing:
        print(f"  ERROR: Team sheet is missing column(s): {', '.join(sorted(missing))}. "
              f"Found: {', '.join(df.columns)}")
        return {}
    team = {}
    for idx, row in df.iterrows():
        name = clean_str(row["Name"])
        if not name or name == "nan":
            continue  # skip blank rows
        try:
            days = float(row["Days Per Week"])
        except (ValueError, TypeError):
            print(f"  WARNING: Team row {idx + 2}: invalid Days Per Week for '{name}', skipping.")
            continue
        if days <= 0:
            print(f"  WARNING: Team row {idx + 2}: '{name}' has {days} days/week (should be positive), skipping.")
            continue
        team[name] = days
    return team


def load_workstreams(filepath):
    """Load workstreams, colors, and priorities from the 'Workstreams' sheet."""
    try:
        df = pd.read_excel(filepath, sheet_name="Workstreams")
    except Exception as e:
        print(f"  WARNING: Could not read Workstreams sheet: {e}")
        return {}
    if df.empty:
        return {}
    df.columns = df.columns.str.strip()
    required = {"Workstream", "Color"}
    missing = required - set(df.columns)
    if missing:
        print(f"  ERROR: Workstreams sheet is missing column(s): {', '.join(sorted(missing))}. "
              f"Found: {', '.join(df.columns)}")
        return {}
    workstreams = {}
    for idx, row in df.iterrows():
        name = clean_str(row["Workstream"])
        if not name or name == "nan":
            continue
        color = clean_str(row["Color"])
        priority = clean_str(row.get("Priority", "")) or "P2"
        if priority not in PRIORITY_VALUES:
            priority = "P2"
        workstreams[name] = {"color": color, "priority": priority}
    return workstreams


def load_tasks(filepath, workstreams=None):
    """Load tasks from the 'Tasks' sheet. Workstreams dict is used for priority inheritance."""
    try:
        df = pd.read_excel(filepath, sheet_name="Tasks")
    except Exception as e:
        print(f"  WARNING: Could not read Tasks sheet: {e}")
        return []
    if df.empty:
        return []
    df.columns = df.columns.str.strip()
    required = {"Task", "Workstream", "Assigned To", "Start Date", "Total Days", "Status"}
    missing = required - set(df.columns)
    if missing:
        print(f"  ERROR: Tasks sheet is missing column(s): {', '.join(sorted(missing))}. "
              f"Found: {', '.join(df.columns)}")
        return []
    tasks = []
    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            task_name = clean_str(row["Task"])
            if not task_name or task_name == "nan":
                continue  # skip blank rows

            start = parse_date(row["Start Date"], context=f"Tasks row {row_num}, 'Start Date'")

            # Total Days (current estimate, supports fractions)
            total_days = float(row["Total Days"])

            # Original Days — auto-fills from Total Days if blank
            original_days = total_days
            if "Original Days" in df.columns and pd.notna(row.get("Original Days")):
                try:
                    original_days = float(row["Original Days"])
                except (ValueError, TypeError):
                    pass  # fallback to total_days

            # Priority — inherits from workstream if blank
            raw_priority = clean_str(row.get("Priority", ""))
            if raw_priority in PRIORITY_VALUES:
                priority = raw_priority
            elif workstreams and clean_str(row["Workstream"]) in workstreams:
                priority = workstreams[clean_str(row["Workstream"])]["priority"]
            else:
                priority = "P2"

            # Status — defaults to Planned if blank
            status = clean_str(row["Status"])
            if not status or status == "nan":
                status = "Planned"

            # Actual End (optional)
            actual_end = None
            if "Actual End" in df.columns and pd.notna(row.get("Actual End")):
                try:
                    actual_end = parse_date(row["Actual End"], context=f"Tasks row {row_num}, 'Actual End'")
                except (ValueError, TypeError):
                    pass

            # Blocked By (optional)
            blocked_by = clean_str(row.get("Blocked By", ""))

            # Deadline (optional)
            deadline = None
            if "Deadline" in df.columns and pd.notna(row.get("Deadline")):
                try:
                    deadline = parse_date(row["Deadline"], context=f"Tasks row {row_num}, 'Deadline'")
                except (ValueError, TypeError):
                    pass

            # Risk/Confidence (optional)
            confidence = clean_str(row.get("Confidence", ""))
            if confidence and confidence not in CONFIDENCE_VALUES:
                confidence = ""

            tasks.append({
                "task": task_name,
                "workstream": clean_str(row["Workstream"]),
                "assigned_to": clean_str(row["Assigned To"]),
                "start_date": start,
                "original_days": original_days,
                "total_days": total_days,
                "priority": priority,
                "status": status,
                "actual_end": actual_end,
                "blocked_by": blocked_by,
                "deadline": deadline,
                "confidence": confidence,
                "notes": clean_str(row.get("Notes", "")),
                "_row": row_num,
            })
        except Exception as e:
            print(f"  WARNING: Could not parse row {row_num}: {e}")
    return tasks


def load_public_holidays(filepath):
    """Load public holidays from the Excel file. Returns set[datetime] (empty if sheet missing)."""
    try:
        df = pd.read_excel(filepath, sheet_name="Public Holidays")
    except (ValueError, Exception):
        # Sheet doesn't exist — backwards compatible
        return set()
    if df.empty:
        return set()
    df.columns = df.columns.str.strip()
    if "Date" not in df.columns:
        print("  WARNING: Public Holidays sheet has no 'Date' column, skipping.")
        return set()

    holidays = set()
    for idx, row in df.iterrows():
        if pd.isna(row["Date"]):
            continue
        try:
            holidays.add(parse_date(row["Date"], context=f"Public Holidays row {idx + 2}, 'Date'"))
        except Exception as e:
            print(f"  WARNING: Could not parse public holiday row {idx + 2}: {e}")
    return holidays


def load_leave(filepath):
    """Load leave entries from the Excel file.
    Returns (leave_dates, leave_entries) where:
      leave_dates = dict[str, set[datetime]] (person -> leave dates for scheduling)
      leave_entries = list[dict] (raw entries with type/dates for console output)
    """
    try:
        df = pd.read_excel(filepath, sheet_name="Leave")
    except (ValueError, Exception):
        # Sheet doesn't exist — backwards compatible
        return {}, []
    if df.empty:
        return {}, []
    df.columns = df.columns.str.strip()
    required = {"Person", "Start Date", "End Date"}
    missing = required - set(df.columns)
    if missing:
        print(f"  WARNING: Leave sheet is missing column(s): {', '.join(sorted(missing))}. Skipping.")
        return {}, []

    leave_dates = {}
    leave_entries = []
    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            person = clean_str(row["Person"])
            if not person or person == "nan":
                continue
            start = parse_date(row["Start Date"], context=f"Leave row {row_num}, 'Start Date'")
            end = parse_date(row["End Date"], context=f"Leave row {row_num}, 'End Date'")
            leave_type = clean_str(row.get("Type", "")) or "Other"
            notes = clean_str(row.get("Notes", ""))

            # Expand date range into individual weekdays
            if person not in leave_dates:
                leave_dates[person] = set()
            current = start
            days_count = 0
            while current <= end:
                if current.weekday() < 5:
                    leave_dates[person].add(norm_date(current))
                    days_count += 1
                current += timedelta(days=1)

            leave_entries.append({
                "person": person,
                "start": start,
                "end": end,
                "type": leave_type,
                "notes": notes,
                "days": days_count,
            })
        except Exception as e:
            print(f"  WARNING: Could not parse leave row {row_num}: {e}")
    return leave_dates, leave_entries


def load_data(filepath):
    """Load all data from the Excel file."""
    team = load_team(filepath)
    workstreams = load_workstreams(filepath)
    tasks = load_tasks(filepath, workstreams=workstreams)
    public_holidays = load_public_holidays(filepath)
    leave_dates, leave_entries = load_leave(filepath)

    # Print load summary for leave/holidays
    if public_holidays:
        print(f"  Public holidays: {len(public_holidays)}")
    if leave_dates:
        total_days = sum(len(dates) for dates in leave_dates.values())
        people = len(leave_dates)
        print(f"  Leave entries: {total_days} day(s) across {people} {'person' if people == 1 else 'people'}")

    return team, workstreams, tasks, public_holidays, leave_dates, leave_entries


# ── Data Validation ──────────────────────────────────────────────────────────

def validate_data(team, workstreams, tasks, public_holidays=None, leave=None):
    """Validate loaded data. Returns (errors, warnings) lists."""
    errors = []
    warnings = []

    if not team:
        errors.append("Team sheet is empty. Add at least one team member.")
    if not workstreams:
        errors.append("Workstreams sheet is empty. Add at least one workstream.")
    if not tasks:
        errors.append("Tasks sheet is empty. Add at least one task.")
        return errors, warnings

    ws_names = set(workstreams.keys())
    team_names = set(team.keys())

    # Validate workstream colours
    hex_re = re.compile(r'^#[0-9A-Fa-f]{6}$')
    for ws_name, ws_info in workstreams.items():
        if not hex_re.match(ws_info["color"]):
            errors.append(f"Workstream '{ws_name}': color '{ws_info['color']}' is not a valid hex code (e.g. #2196F3).")
        if ws_info["priority"] not in PRIORITY_VALUES:
            warnings.append(f"Workstream '{ws_name}': priority '{ws_info['priority']}' not recognised.")

    # Track tasks to skip (non-positive days) via _skip flag

    for task in tasks:
        row = task.get("_row", "?")

        if not task["task"] or task["task"] == "nan":
            errors.append(f"Row {row}: Task name is empty.")

        if task["workstream"] not in ws_names:
            close = difflib.get_close_matches(task["workstream"], list(ws_names), n=1, cutoff=0.4)
            hint = f" Did you mean: '{close[0]}'?" if close else ""
            errors.append(f"Row {row}: Workstream '{task['workstream']}' not found in Workstreams sheet.{hint}")

        if task["assigned_to"] not in team_names:
            errors.append(f"Row {row}: '{task['assigned_to']}' not in Team sheet. "
                          f"Team members: {', '.join(team_names)}")

        if not isinstance(task["start_date"], datetime):
            errors.append(f"Row {row}: Invalid start date '{task['start_date']}'. Use YYYY-MM-DD format.")

        if task["total_days"] <= 0:
            errors.append(f"Row {row}: Task '{task['task']}' has {task['total_days']} days "
                          f"(must be positive). Skipping this task.")
            task["_skip"] = True

        if task["status"] not in STATUS_VALUES:
            warnings.append(f"Row {row}: Status '{task['status']}' not recognised. "
                            f"Valid: {', '.join(STATUS_VALUES)}")

        if task["priority"] not in PRIORITY_VALUES:
            warnings.append(f"Row {row}: Priority '{task['priority']}' not recognised. "
                            f"Valid: {', '.join(PRIORITY_VALUES)}")

        # Validate workstream priorities
        if task["workstream"] in workstreams:
            ws_priority = workstreams[task["workstream"]]["priority"]
            if priority_sort_key(task["priority"]) < priority_sort_key(ws_priority):
                warnings.append(
                    f"Row {row}: Task priority {task['priority']} is higher than its "
                    f"workstream priority {ws_priority} ({task['workstream']})")

        # Validate original_days
        if task["original_days"] <= 0:
            warnings.append(f"Row {row}: Task '{task['task']}' has invalid Original Days ({task['original_days']}).")

        # Validate deadline
        if task.get("deadline") and isinstance(task["start_date"], datetime):
            if task["deadline"] < task["start_date"]:
                warnings.append(f"Row {row}: Deadline ({task['deadline'].strftime('%Y-%m-%d')}) is before start date.")

    # Remove tasks with non-positive days (they'd crash the scheduler)
    tasks[:] = [t for t in tasks if not t.get("_skip")]

    # Validate leave entries
    if leave:
        for person, dates in leave.items():
            if person not in team_names:
                warnings.append(f"Leave: '{person}' not found in Team sheet. "
                                f"Team members: {', '.join(team_names)}")

    # Warn about public holidays on weekends
    if public_holidays:
        for hol in sorted(public_holidays):
            if hol.weekday() >= 5:
                warnings.append(f"Public holiday {hol.strftime('%Y-%m-%d')} falls on a weekend (has no effect).")

    return errors, warnings


# ── Schedule Calculation ─────────────────────────────────────────────────────

def get_end_date(start_date, total_working_days, public_holidays=None, person_leave=None):
    """Calculate end date by adding working days (supports fractional days).
    Skips weekends, public holidays, and person leave days.
    Returns (end_date, working_days_list, day_allocations_dict)."""
    start_date = norm_date(start_date)
    if total_working_days <= 0:
        return start_date, [], {}

    current = start_date
    remaining = total_working_days
    working_days = []
    day_allocations = {}

    while remaining > 0:
        if is_working_day(current, public_holidays, person_leave):
            alloc = min(remaining, 1.0)
            working_days.append(current)
            day_allocations[current] = alloc
            remaining -= alloc
            if remaining <= 0:
                return current, working_days, day_allocations
        current += timedelta(days=1)

    return current, working_days, day_allocations


def calculate_schedule(tasks, public_holidays=None, leave=None):
    """For each task, compute start/end dates, working days, and day allocations."""
    for task in tasks:
        start = norm_date(task["start_date"])
        person_leave = leave.get(task["assigned_to"]) if leave else None

        # Snap to next working day if start falls on non-working day
        while not is_working_day(start, public_holidays, person_leave):
            start += timedelta(days=1)
        task["start_date"] = start

        end_date, working_days, day_allocations = get_end_date(
            start, task["total_days"], public_holidays, person_leave
        )
        task["end_date"] = end_date
        task["working_days"] = working_days
        task["day_allocations"] = day_allocations

        # Compute actual end date info for Complete tasks with Actual End
        if task["status"] == "Complete" and task["actual_end"]:
            ae = task["actual_end"]
            while not is_working_day(ae, public_holidays, person_leave):
                ae -= timedelta(days=1)
            task["actual_end_date"] = ae

            # Count working days between start and actual end
            actual_wd = 0
            d = start
            while d <= ae:
                if is_working_day(d, public_holidays, person_leave):
                    actual_wd += 1
                d += timedelta(days=1)
            task["actual_working_days"] = actual_wd
        else:
            task["actual_end_date"] = None
            task["actual_working_days"] = None

    return tasks


# ── Capacity Calculation ─────────────────────────────────────────────────────

def get_week_start(date):
    """Get the Monday of the week containing the given date."""
    return date - timedelta(days=date.weekday())


def calculate_capacity(tasks, team, public_holidays=None, leave=None):
    """Calculate per-person per-week allocation and leave-adjusted available capacity.
    Returns (allocation, weeks, available) where available = {week: {person: adjusted_days}}."""
    if not tasks:
        return {}, [], {}

    all_dates = []
    for t in tasks:
        all_dates.extend(t.get("working_days", []))
    if not all_dates:
        return {}, [], {}

    min_date = get_week_start(min(all_dates))
    max_week_start = get_week_start(max(all_dates))

    weeks = []
    current = min_date
    while current <= max_week_start:
        weeks.append(current)
        current += timedelta(days=7)

    allocation = {w: {name: 0.0 for name in team} for w in weeks}

    # Calculate per-person per-week available capacity (adjusted for holidays + leave)
    available = {}
    for w in weeks:
        available[w] = {}
        for name, days_pw in team.items():
            person_leave = leave.get(name, set()) if leave else set()
            working_days_this_week = 0
            for offset in range(5):  # Mon-Fri
                day = w + timedelta(days=offset)
                if is_working_day(day, public_holidays, person_leave):
                    working_days_this_week += 1
            # Scale by days_per_week / 5 for part-time
            available[w][name] = (days_pw / 5) * working_days_this_week

    for task in tasks:
        person = task["assigned_to"]
        if person not in team:
            continue
        for day, alloc in task.get("day_allocations", {}).items():
            ws = get_week_start(day)
            if ws in allocation:
                allocation[ws][person] += alloc

    return allocation, weeks, available


def calculate_monthly_capacity(tasks, team, public_holidays=None, leave=None):
    """Calculate per-person per-month allocation and leave-adjusted available capacity."""
    if not tasks:
        return {}, [], {}

    all_dates = []
    for t in tasks:
        all_dates.extend(t.get("working_days", []))
    if not all_dates:
        return {}, [], {}

    min_date = min(all_dates)
    max_date = max(all_dates)

    months = []
    current = datetime(min_date.year, min_date.month, 1)
    end = datetime(max_date.year, max_date.month, 1)
    while current <= end:
        months.append(current)
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    allocation = {m: {name: 0.0 for name in team} for m in months}
    available = {}
    for m in months:
        wd = working_days_in_month(m.year, m.month, public_holidays)
        available[m] = {}
        for name, days_pw in team.items():
            # Count person's leave days in this month
            person_leave = leave.get(name, set()) if leave else set()
            leave_days_in_month = 0
            if person_leave:
                _, num_days = monthrange(m.year, m.month)
                for d in range(1, num_days + 1):
                    dt = datetime(m.year, m.month, d)
                    if dt.weekday() < 5 and dt in person_leave:
                        leave_days_in_month += 1
            # /5 = calendar weekdays (Mon-Fri). Scales part-time correctly:
            # e.g. 3 days/week person in a 22-workday month → (3/5)*22 = 13.2 available days
            # Leave days are subtracted after scaling
            available[m][name] = (days_pw / 5) * wd - (days_pw / 5) * leave_days_in_month

    for task in tasks:
        person = task["assigned_to"]
        if person not in team:
            continue
        for day, alloc in task.get("day_allocations", {}).items():
            month_key = datetime(day.year, day.month, 1)
            if month_key in allocation:
                allocation[month_key][person] += alloc

    return allocation, months, available


# ── Roadmap Helpers ──────────────────────────────────────────────────────────

def aggregate_workstreams(tasks, workstreams):
    """Aggregate task data per workstream for roadmap view."""
    ws_data = {}
    for ws_name in workstreams:
        ws_tasks = [t for t in tasks if t["workstream"] == ws_name]
        if not ws_tasks:
            continue
        earliest_start = min(t["start_date"] for t in ws_tasks)
        latest_end = max(t["end_date"] for t in ws_tasks)

        day_counts = {}
        for t in ws_tasks:
            for wd in t.get("working_days", []):
                day_counts[wd] = day_counts.get(wd, 0) + 1

        task_starts = [(t["start_date"], t["task"]) for t in ws_tasks]
        has_blocked = any(t["status"] == "On Hold" for t in ws_tasks)
        blocked_tasks = [t for t in ws_tasks if t["status"] == "On Hold"]

        ws_data[ws_name] = {
            "start": earliest_start,
            "end": latest_end,
            "task_count": len(ws_tasks),
            "day_counts": day_counts,
            "max_concurrent": max(day_counts.values()) if day_counts else 1,
            "task_starts": task_starts,
            "has_blocked": has_blocked,
            "blocked_tasks": blocked_tasks,
        }
    return ws_data


def get_quarter_boundaries(start_date, end_date):
    """Return a list of quarter start dates within the range."""
    quarter_months = [1, 4, 7, 10]
    boundaries = []
    year = start_date.year
    while year <= end_date.year + 1:
        for m in quarter_months:
            d = datetime(year, m, 1)
            if start_date - timedelta(days=30) <= d <= end_date + timedelta(days=30):
                boundaries.append(d)
        year += 1
    return sorted(boundaries)


def get_quarter_label(date):
    """Return 'Q1 2026' style label for a date."""
    quarter = (date.month - 1) // 3 + 1
    return f"Q{quarter} {date.year}"


# ── Schedule Suggestions ─────────────────────────────────────────────────────

def print_schedule_suggestions(tasks, team, allocation, weeks, available=None,
                               public_holidays=None, leave=None):
    """Analyse the schedule and print actionable suggestions."""
    suggestions = []
    today = datetime.now()

    # Group tasks by person, sorted by start date
    person_tasks = {}
    for t in tasks:
        person_tasks.setdefault(t["assigned_to"], []).append(t)
    for p in person_tasks:
        person_tasks[p].sort(key=lambda t: t["start_date"])

    # 1. Early finishers
    for t in tasks:
        if t["status"] == "Complete" and t.get("actual_end_date"):
            planned_end = t["end_date"]
            actual_end = t["actual_end_date"]
            if actual_end < planned_end:
                person = t["assigned_to"]
                person_leave = leave.get(person, set()) if leave else None
                days_early = count_working_days(
                    actual_end + timedelta(days=1), planned_end,
                    public_holidays, person_leave)
                if days_early > 0:
                    subsequent = [st for st in person_tasks.get(person, [])
                                  if st["start_date"] > t["start_date"]
                                  and st["status"] in ("Planned", "In Progress")
                                  and st["task"] != t["task"]]
                    if subsequent:
                        next_task = subsequent[0]
                        new_start = actual_end + timedelta(days=1)
                        while not is_working_day(new_start, public_holidays, person_leave):
                            new_start += timedelta(days=1)
                        suggestions.append(
                            f"  {t['task']} finished {days_early} day{'s' if days_early != 1 else ''} early.\n"
                            f"    -> {next_task['task']} ({person}) could start "
                            f"{new_start.strftime('%d %b')} instead of {next_task['start_date'].strftime('%d %b')}")

    # 2. Overdue tasks
    for t in tasks:
        if t["status"] == "In Progress" and t["end_date"] < today:
            person_leave = leave.get(t["assigned_to"], set()) if leave else None
            overdue_days = count_working_days(
                t["end_date"] + timedelta(days=1), today,
                public_holidays, person_leave)
            if overdue_days > 0:
                suggestions.append(
                    f"  WARNING: {t['task']} is {overdue_days} day{'s' if overdue_days != 1 else ''} "
                    f"overdue (planned end: {t['end_date'].strftime('%d %b')})")

    # 3. Blocked duration
    for t in tasks:
        if t["status"] == "On Hold":
            person_leave = leave.get(t["assigned_to"], set()) if leave else None
            blocked_days = count_working_days(
                t["start_date"], today, public_holidays, person_leave)
            msg = f"  {t['task']} has been on hold for {blocked_days} working day{'s' if blocked_days != 1 else ''}"
            if t["blocked_by"]:
                msg += f"\n    Blocked by: {t['blocked_by']}"
            suggestions.append(msg)

    # 4. Leave overlaps — warn when a person has leave during an active task
    if leave:
        for t in tasks:
            if t["status"] in ("Planned", "In Progress"):
                person = t["assigned_to"]
                person_leave = leave.get(person, set())
                if person_leave:
                    t_start = norm_date(t["start_date"])
                    t_end = norm_date(t["end_date"])
                    overlap = sorted(d for d in person_leave
                                     if t_start <= norm_date(d) <= t_end)
                    if overlap:
                        suggestions.append(
                            f"  {person} has {len(overlap)} leave day(s) during '{t['task']}' "
                            f"({overlap[0].strftime('%d %b')} - {overlap[-1].strftime('%d %b')})")

    # 5. Capacity gaps (capped at 5 to avoid verbose output on long timelines)
    max_gap_entries = 5
    gap_count = 0
    gap_overflow = 0
    if weeks:
        for w in weeks:
            if w < today:
                continue  # skip past weeks
            for person in team:
                person_alloc = allocation[w].get(person, 0)
                person_avail = available[w][person] if available and w in available else team[person]
                free = person_avail - person_alloc
                if free >= 3:  # significant spare capacity
                    if gap_count < max_gap_entries:
                        suggestions.append(
                            f"  {person} has spare capacity in w/c {w.strftime('%d %b')} "
                            f"({person_alloc:.1f} day{'s' if person_alloc != 1 else ''} allocated, "
                            f"{free:.1f} day{'s' if free != 1 else ''} free)")
                        gap_count += 1
                    else:
                        gap_overflow += 1
        if gap_overflow > 0:
            suggestions.append(f"  ... and {gap_overflow} more spare capacity gap{'s' if gap_overflow != 1 else ''} not shown")

    if suggestions:
        print()
        print("SCHEDULE SUGGESTIONS:")
        for s in suggestions:
            print(s)
        print()


# ── Chart: Gantt + Weekly Capacity ───────────────────────────────────────────

def _draw_weekend_shading(ax, date_min, date_max):
    """Draw light grey vertical bands for weekend days on a date-axis chart."""
    d = date_min
    while d <= date_max:
        if d.weekday() == 5:  # Saturday
            sat_num = mdates.date2num(d)
            sun_num = sat_num + 1
            ax.axvspan(sat_num, sun_num + 1, color="#E0E0E0", alpha=0.15, zorder=0)
        d += timedelta(days=1)


def render_gantt(tasks, team, workstreams, weeks, output_path,
                 public_holidays=None, leave=None):
    """Render the standalone Gantt chart."""
    apply_style()

    # Sort workstreams by priority then original Excel order
    ws_order_items = list(workstreams.items())
    ws_order_items.sort(key=lambda item: (priority_sort_key(item[1]["priority"]),
                                          list(workstreams.keys()).index(item[0])))
    ws_order = [name for name, _ in ws_order_items]

    # Group tasks by workstream
    grouped = {}
    for ws_name in ws_order:
        ws_tasks = [t for t in tasks if t["workstream"] == ws_name]
        if ws_tasks:
            # Sort tasks within workstream: by task priority, then start date
            ws_tasks.sort(key=lambda t: (priority_sort_key(t["priority"]), t["start_date"], t.get("_row", 0)))
            grouped[ws_name] = ws_tasks

    if not grouped or not weeks:
        print("  No Gantt data. Check: tasks have positive Total Days, valid Start Date, and are not all filtered out.")
        return

    # Count rows
    total_rows = sum(len(ts) + 1 for ts in grouped.values())
    person_list = list(team.keys())
    person_hatch = {name: PERSON_HATCHES.get(i, "") for i, name in enumerate(person_list)}

    # ── Figure (single Gantt axis, no capacity panel) ──
    fig_height = max(8, total_rows * 0.5 + 3)
    fig = plt.figure(figsize=(STYLE["fig_width"], fig_height), facecolor=STYLE["bg_color"])
    ax_gantt = fig.add_axes([0.18, 0.10, 0.77, 0.80])

    date_min = weeks[0] - timedelta(days=3)
    date_max = weeks[-1] + timedelta(days=9)

    # ── Alternating row shading ──
    for i in range(total_rows):
        shade = STYLE["row_shade_even"] if i % 2 == 0 else STYLE["row_shade_odd"]
        ax_gantt.axhspan(i - 0.5, i + 0.5, color=shade, alpha=0.6, zorder=0)

    # ── Weekend shading ──
    _draw_weekend_shading(ax_gantt, date_min, date_max)

    # ── Render Gantt bars ──
    y_pos = total_rows - 1
    y_ticks = []
    y_labels = []
    y_colors = []

    for ws_name in ws_order:
        if ws_name not in grouped:
            continue
        ws_info = workstreams[ws_name]
        ws_color = ws_info["color"]
        ws_priority = ws_info["priority"]
        ws_tasks = grouped[ws_name]

        ws_pstyle = PRIORITY_STYLES.get(ws_priority, PRIORITY_STYLES["P2"])

        # Workstream header: accent bar + subtle background
        ax_gantt.axhspan(y_pos - 0.4, y_pos + 0.4,
                         color=ws_color, alpha=STYLE["header_bg_alpha"], zorder=0)
        accent_width = (mdates.date2num(date_max) - mdates.date2num(date_min)) * 0.008
        ax_gantt.barh(y_pos, accent_width,
                      left=mdates.date2num(date_min),
                      height=0.8, color=ws_color, alpha=0.9,
                      edgecolor="none", zorder=2)

        # Priority badge on workstream header
        header_label = f"{ws_priority}  {ws_name}"

        y_ticks.append(y_pos)
        y_labels.append(header_label)
        y_colors.append(ws_color)
        y_pos -= 1

        # Task bars
        for task in ws_tasks:
            start_num = mdates.date2num(task["start_date"])
            end_num = mdates.date2num(task["end_date"])
            duration = max(end_num - start_num + 1, 1)

            t_priority = task["priority"]
            t_pstyle = PRIORITY_STYLES.get(t_priority, PRIORITY_STYLES["P2"])

            alpha = t_pstyle["alpha"]
            bar_color = ws_color
            edge = ws_color
            lw = t_pstyle["linewidth"]
            hatch = person_hatch.get(task["assigned_to"], "")
            bar_linestyle = "-"

            is_complete = task["status"] == "Complete"
            is_on_hold = task["status"] == "On Hold"
            is_planned = task["status"] == "Planned"

            if is_complete:
                alpha = 0.35
                edge = "#AAAAAA"
                lw = 1.0
            elif is_on_hold:
                bar_color = STYLE["on_hold_color"]
                alpha = 0.75
                lw = 2.0
                edge = STYLE["on_hold_edge_color"]
                hatch = "xx"  # cross-hatch for blocked
                bar_linestyle = "--"
            elif is_planned:
                alpha = t_pstyle["alpha"] * STYLE["planned_alpha_factor"]

            # ── Planned vs Actual for Complete tasks with Actual End ──
            if is_complete and task.get("actual_end_date"):
                actual_end_num = mdates.date2num(task["actual_end_date"])
                actual_duration = max(actual_end_num - start_num + 1, 1)

                # Ghost bar: planned position (dashed outline, no fill)
                draw_rounded_bar(ax_gantt, start_num, y_pos, duration,
                                 STYLE["bar_height"], "none",
                                 alpha=0.6, edgecolor="#999999",
                                 linewidth=1.0, linestyle="--", zorder=2)

                # Solid bar: actual position
                draw_rounded_bar(ax_gantt, start_num, y_pos, actual_duration,
                                 STYLE["bar_height"], ws_color,
                                 alpha=0.45, edgecolor=ws_color,
                                 linewidth=1.2, hatch=hatch, zorder=3)

                # Variance label
                planned_wd = len(task.get("working_days", []))
                actual_wd = task.get("actual_working_days", planned_wd)
                diff = actual_wd - planned_wd
                if diff > 0:
                    var_label = f"+{diff}d late"
                    var_color = STYLE["late_color"]
                elif diff < 0:
                    var_label = f"{diff}d early"
                    var_color = STYLE["early_color"]
                else:
                    var_label = "on time"
                    var_color = STYLE["text_muted"]

                ax_gantt.text(max(start_num + actual_duration, start_num + duration) + 0.5,
                              y_pos + 0.15, var_label,
                              fontsize=STYLE["small_size"] - 0.5,
                              color=var_color, fontweight="bold",
                              va="center", ha="left", zorder=6)
            else:
                # Standard bar
                draw_rounded_bar(ax_gantt, start_num, y_pos, duration,
                                 STYLE["bar_height"], bar_color,
                                 alpha=alpha, edgecolor=edge,
                                 linewidth=lw, hatch=hatch,
                                 linestyle=bar_linestyle)

            # ── Build label text ──
            sym = STATUS_SYMBOLS.get(task["status"], "")
            task_name_display = task["task"]
            if len(task_name_display) > 40:
                task_name_display = task_name_display[:37] + "..."
            label_parts = [f"{sym} {task_name_display} ({task['assigned_to']}) {task['total_days']:.4g} wd"]

            # Estimation drift
            drift_text = ""
            if task["original_days"] != task["total_days"] and task["original_days"] > 0:
                pct_change = ((task["total_days"] - task["original_days"]) / task["original_days"]) * 100
                sign = "+" if pct_change > 0 else ""
                drift_text = f"(was {task['original_days']:.4g}d, now {task['total_days']:.4g}d {sign}{pct_change:.0f}%)"
                label_parts.append(drift_text)

            # Blocked info (no artificial cap on blocked duration)
            if is_on_hold:
                today = datetime.now()
                person_leave_set = leave.get(task["assigned_to"], set()) if leave else None
                blocked_days = count_working_days(
                    task["start_date"], today, public_holidays, person_leave_set)
                blocked_label = f"({blocked_days}d blocked)"
                if task["blocked_by"]:
                    blocked_label = f"On Hold: {task['blocked_by']} ({blocked_days}d)"
                label_parts.append(blocked_label)

            label_text = " ".join(label_parts)

            # Smart label placement
            label_weight = t_pstyle["label_weight"]
            label_size = STYLE["small_size"] + t_pstyle["label_size_bump"]

            # Drift color for label
            drift_color = None
            if task["original_days"] != task["total_days"] and task["original_days"] > 0:
                if task["total_days"] > task["original_days"]:
                    drift_color = STYLE["drift_increase_color"]
                else:
                    drift_color = STYLE["drift_decrease_color"]

            estimated_label_width = len(label_text) * 0.65
            if duration > estimated_label_width and not is_complete:
                # Label fits inside bar
                ax_gantt.text(
                    start_num + duration / 2, y_pos, label_text,
                    va="center", ha="center",
                    fontsize=label_size, fontweight=label_weight,
                    color="white",
                    path_effects=[pe.withStroke(linewidth=2.5, foreground=ws_color)],
                    clip_on=True, zorder=5,
                )
            else:
                # Label overflows to the right
                label_color = STYLE["text_muted"] if is_complete else STYLE["text_primary"]
                bar_end = start_num + duration
                if is_complete and task.get("actual_end_date"):
                    actual_end_num = mdates.date2num(task["actual_end_date"])
                    bar_end = max(bar_end, start_num + max(actual_end_num - start_num + 1, 1))

                ax_gantt.text(
                    bar_end + 0.5, y_pos, label_text,
                    va="center", ha="left",
                    fontsize=label_size, fontweight=label_weight,
                    color=label_color, clip_on=True, zorder=5,
                )

                # Coloured drift text (separate, positioned after main label)
                if drift_text and drift_color:
                    # Calculate offset for drift text
                    drift_x = bar_end + 0.5 + len(label_text.replace(drift_text, "")) * 0.55
                    ax_gantt.text(
                        drift_x, y_pos - 0.22, drift_text,
                        va="center", ha="left",
                        fontsize=label_size - 0.5, fontweight="bold",
                        color=drift_color, clip_on=True, zorder=5,
                    )

            # ── Confidence dot ──
            if task.get("confidence") and task["confidence"] in CONFIDENCE_COLORS:
                conf_color = CONFIDENCE_COLORS[task["confidence"]]
                ax_gantt.plot(start_num - 0.8, y_pos, "o",
                              color=conf_color, markersize=5, zorder=6,
                              markeredgecolor="white", markeredgewidth=0.5)

            # ── Deadline marker ──
            if task.get("deadline"):
                deadline_num = mdates.date2num(task["deadline"])
                if mdates.date2num(date_min) <= deadline_num <= mdates.date2num(date_max):
                    ax_gantt.plot(deadline_num, y_pos, "D",
                                  color="#D32F2F", markersize=6, zorder=7,
                                  markeredgecolor="white", markeredgewidth=0.5)
                    # Red tint on overshoot portion
                    if task.get("end_date") and task["end_date"] > task["deadline"]:
                        overshoot_start = deadline_num
                        overshoot_end = mdates.date2num(task["end_date"]) + 1
                        draw_rounded_bar(ax_gantt, overshoot_start, y_pos,
                                         overshoot_end - overshoot_start,
                                         STYLE["bar_height"], "#D32F2F",
                                         alpha=0.2, edgecolor="none",
                                         linewidth=0, zorder=4)

            y_ticks.append(y_pos)
            y_labels.append("")
            y_colors.append(ws_color)
            y_pos -= 1

    # Gantt axis formatting
    ax_gantt.set_yticks(y_ticks)
    ax_gantt.set_yticklabels(y_labels, fontsize=STYLE["tick_size"])
    for tick_label, color in zip(ax_gantt.get_yticklabels(), y_colors):
        if tick_label.get_text():
            tick_label.set_fontweight("bold")
            tick_label.set_fontsize(STYLE["label_size"])
            tick_label.set_color(color)

    ax_gantt.set_xlim(mdates.date2num(date_min), mdates.date2num(date_max))
    ax_gantt.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=0))
    ax_gantt.xaxis.set_major_formatter(mdates.DateFormatter("%d %b"))
    plt.setp(ax_gantt.xaxis.get_majorticklabels(), rotation=45, ha="right",
             fontsize=STYLE["tick_size"])

    draw_today_line(ax_gantt, date_min, date_max, total_rows - 0.5)

    # ── Public holiday dotted lines ──
    if public_holidays:
        for hol in sorted(public_holidays):
            hol_num = mdates.date2num(hol)
            if mdates.date2num(date_min) <= hol_num <= mdates.date2num(date_max):
                ax_gantt.axvline(hol_num, color=STYLE["holiday_edge_color"],
                                 linewidth=0.8, linestyle=":", alpha=0.4, zorder=1)

    # ── Per-person leave markers (▼) on task rows ──
    if leave:
        # Build a map of y_pos per task for leave markers
        leave_y = total_rows - 1
        for ws_name in ws_order:
            if ws_name not in grouped:
                continue
            leave_y -= 1  # skip header row
            for task in grouped[ws_name]:
                person = task["assigned_to"]
                person_leave = leave.get(person, set())
                if person_leave:
                    for d in sorted(person_leave):
                        d_num = mdates.date2num(d)
                        s_num = mdates.date2num(task["start_date"])
                        e_num = mdates.date2num(task["end_date"])
                        if s_num <= d_num <= e_num:
                            ax_gantt.plot(d_num, leave_y + 0.35, marker="v",
                                          color=STYLE["leave_edge_color"],
                                          markersize=4, alpha=0.7, zorder=6)
                leave_y -= 1

    style_axes(ax_gantt, title="Gantt View", show_grid_x=True)

    # ── Legend ──
    legend_handles = []
    for ws_name in ws_order:
        if ws_name in grouped:
            ws_color = workstreams[ws_name]["color"]
            legend_handles.append(mpatches.Patch(facecolor=ws_color, edgecolor=ws_color, label=ws_name))
    for pidx, person in enumerate(person_list):
        legend_handles.append(mpatches.Patch(
            facecolor="#CCCCCC", edgecolor="#333333",
            hatch=PERSON_HATCHES.get(pidx, ""), label=f"{person}"
        ))
    # Priority legend entries
    for p in PRIORITY_VALUES:
        ps = PRIORITY_STYLES[p]
        legend_handles.append(mpatches.Patch(
            facecolor="#888888", edgecolor="#333333",
            alpha=ps["alpha"], linewidth=ps["linewidth"],
            label=f"{p}",
        ))
    # Status legend entries
    legend_handles.append(mpatches.Patch(
        facecolor=STYLE["on_hold_color"], edgecolor=STYLE["on_hold_edge_color"],
        alpha=0.75, linewidth=2.0, hatch="xx", linestyle="--",
        label="On Hold",
    ))
    legend_handles.append(mpatches.Patch(
        facecolor="#888888", edgecolor="#AAAAAA",
        alpha=0.35, linewidth=1.0,
        label="Complete",
    ))
    if public_holidays:
        legend_handles.append(plt.Line2D([0], [0], color=STYLE["holiday_edge_color"],
                                          linewidth=0.8, linestyle=":", alpha=0.6,
                                          label="Public holiday"))
    if leave:
        legend_handles.append(plt.Line2D([0], [0], marker="v", color="w",
                                          markerfacecolor=STYLE["leave_edge_color"],
                                          markersize=6, alpha=0.7,
                                          label="Leave day"))
    # Deadline marker in legend (if any task has a deadline)
    if any(t.get("deadline") for t in tasks):
        legend_handles.append(plt.Line2D([0], [0], marker="D", color="w",
                                          markerfacecolor="#D32F2F",
                                          markersize=6, markeredgecolor="white",
                                          label="Deadline"))
    # Confidence dots in legend (if any task has confidence)
    if any(t.get("confidence") for t in tasks):
        for conf, color in CONFIDENCE_COLORS.items():
            if any(t.get("confidence") == conf for t in tasks):
                legend_handles.append(plt.Line2D([0], [0], marker="o", color="w",
                                                  markerfacecolor=color,
                                                  markersize=5,
                                                  label=f"{conf} confidence"))

    ax_gantt.legend(
        handles=legend_handles, loc="upper center",
        bbox_to_anchor=(0.5, -0.06), ncol=min(5, max(2, len(legend_handles) // 2)),
        fontsize=STYLE["small_size"], frameon=True,
        framealpha=0.95, edgecolor=STYLE["grid_color"],
        fancybox=True, columnspacing=1.5, handletextpad=0.5,
    )

    # Header & footer
    date_range = f"{weeks[0].strftime('%d %b %Y')} \u2014 {weeks[-1].strftime('%d %b %Y')}"
    add_header_footer(fig, f"Gantt Chart: {date_range}")

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fig.savefig(output_path, dpi=STYLE["dpi"], bbox_inches="tight", facecolor=STYLE["bg_color"])
    plt.close(fig)
    print(f"  Gantt chart saved: {output_path}")


# ── Chart: Weekly Capacity ──────────────────────────────────────────────────

def render_weekly(tasks, team, workstreams, allocation, weeks, available,
                  output_path, public_holidays=None, leave=None):
    """Render the standalone weekly capacity chart with per-person side-by-side bars."""
    apply_style()

    if not weeks:
        print("  No weekly data. Check: tasks have positive Total Days and valid Start Date.")
        return

    person_list = list(team.keys())
    n_persons = len(person_list)
    x_positions = np.arange(len(weeks))

    # Per-person allocation and available per week
    person_bars = {}
    person_avail = {}
    for person in person_list:
        person_bars[person] = [allocation[w].get(person, 0) for w in weeks]
        person_avail[person] = [available[w][person] if w in available else team[person]
                                for w in weeks]

    # Figure
    fig_height = max(6, n_persons * 1.5 + 3)
    fig = plt.figure(figsize=(STYLE["fig_width"], fig_height), facecolor=STYLE["bg_color"])
    ax = fig.add_axes([0.08, 0.15, 0.88, 0.72])

    bar_group_width = 0.8
    bar_width = bar_group_width / n_persons

    # Holiday week shading
    if public_holidays:
        for i, w in enumerate(weeks):
            hols_this_week = sum(1 for offset in range(5)
                                 if datetime(w.year, w.month, w.day) + timedelta(days=offset)
                                 in public_holidays)
            if hols_this_week > 0:
                ax.axvspan(i - 0.45, i + 0.45, color=STYLE["holiday_color"],
                           alpha=0.3, zorder=0)
                ax.text(i, -0.5, f"{hols_this_week} hol",
                        ha="center", va="top", fontsize=5.5,
                        color=STYLE["holiday_edge_color"], fontstyle="italic")

    # Draw per-person bars side by side
    for pidx, person in enumerate(person_list):
        bar_x = x_positions - bar_group_width / 2 + bar_width * pidx + bar_width / 2
        values = np.array(person_bars[person])
        avail_values = np.array(person_avail[person])

        colors = []
        for i in range(len(weeks)):
            if avail_values[i] > 0 and values[i] > avail_values[i]:
                colors.append(STYLE["over_capacity_color"])
            else:
                colors.append(STYLE["under_capacity_colors"][pidx % len(STYLE["under_capacity_colors"])])

        ax.bar(
            bar_x, values, bar_width * 0.9,
            color=colors, alpha=0.85,
            edgecolor="white", linewidth=0.5,
            hatch=PERSON_HATCHES.get(pidx, ""),
            label=person, zorder=3,
        )

        # Per-person capacity line (dashed)
        ax.plot(bar_x, avail_values,
                color=STYLE["under_capacity_colors"][pidx % len(STYLE["under_capacity_colors"])],
                linewidth=1.5, linestyle="--", alpha=0.6, zorder=4,
                marker=".", markersize=3)

        # Utilisation % labels per person
        for i in range(len(weeks)):
            if avail_values[i] > 0:
                pct = (values[i] / avail_values[i]) * 100
                if values[i] > 0.1:  # Only label non-trivial allocations
                    label_color = STYLE["over_capacity_color"] if pct > 100 else STYLE["text_secondary"]
                    weight = "bold" if pct > 100 else "normal"
                    ax.text(bar_x[i], values[i] + 0.15, f"{pct:.0f}%",
                            ha="center", fontsize=5.5, color=label_color,
                            fontweight=weight, zorder=5)
                    # Over-capacity annotation: show overshoot amount
                    if pct > 100:
                        overshoot = values[i] - avail_values[i]
                        ax.text(bar_x[i], values[i] + 0.55, f"+{overshoot:.1f}d",
                                ha="center", fontsize=4.5, color=STYLE["over_capacity_color"],
                                fontweight="bold", zorder=5)

        # Leave markers
        if leave and person in leave:
            person_leave_dates = leave[person]
            for i, w in enumerate(weeks):
                leave_days_this_week = sum(
                    1 for offset in range(5)
                    if datetime(w.year, w.month, w.day) + timedelta(days=offset) in person_leave_dates
                )
                if leave_days_this_week > 0:
                    ax.text(bar_x[i], -0.3, f"{leave_days_this_week}L",
                            ha="center", va="top", fontsize=5,
                            color=STYLE["leave_edge_color"], fontweight="bold")

    # X axis
    week_labels = [w.strftime("%d %b") for w in weeks]
    ax.set_xticks(x_positions)
    ax.set_xticklabels(week_labels, rotation=45, ha="right", fontsize=STYLE["tick_size"])

    # Y axis
    max_alloc = max(max(person_bars[p]) for p in person_list) if person_list else 5
    max_avail = max(max(person_avail[p]) for p in person_list) if person_list else 5
    ax.set_ylim(-0.8, max(max_alloc, max_avail) * 1.25)

    # Legend
    legend_handles = []
    for pidx, person in enumerate(person_list):
        color = STYLE["under_capacity_colors"][pidx % len(STYLE["under_capacity_colors"])]
        legend_handles.append(mpatches.Patch(
            facecolor=color, edgecolor="white",
            hatch=PERSON_HATCHES.get(pidx, ""), alpha=0.85,
            label=f"{person} ({team[person]:.4g}d/wk)"
        ))
    legend_handles.append(mpatches.Patch(
        facecolor=STYLE["over_capacity_color"], edgecolor="white",
        alpha=0.85, label="Over capacity"
    ))
    if public_holidays:
        legend_handles.append(mpatches.Patch(
            facecolor=STYLE["holiday_color"], edgecolor=STYLE["holiday_edge_color"],
            alpha=0.3, label="Public holiday week"
        ))
    if leave:
        legend_handles.append(mpatches.Patch(
            facecolor=STYLE["leave_color"], edgecolor=STYLE["leave_edge_color"],
            alpha=0.8, label="Leave (NL marker)"
        ))

    ax.legend(handles=legend_handles, loc="upper right",
              fontsize=STYLE["small_size"], framealpha=0.9,
              edgecolor=STYLE["grid_color"], fancybox=True)

    style_axes(ax, title="Weekly Capacity Utilisation (Per Person)",
               ylabel="Days Allocated", show_grid_y=True)

    # Header & footer
    date_range = f"{weeks[0].strftime('%d %b %Y')} \u2014 {weeks[-1].strftime('%d %b %Y')}"
    add_header_footer(fig, f"Weekly Capacity: {date_range}")

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fig.savefig(output_path, dpi=STYLE["dpi"], bbox_inches="tight", facecolor=STYLE["bg_color"])
    plt.close(fig)
    print(f"  Weekly capacity chart saved: {output_path}")


# ── Chart: Monthly Capacity ─────────────────────────────────────────────────

def render_monthly_capacity(tasks, team, workstreams, output_path,
                            public_holidays=None, leave=None):
    """Render the monthly capacity utilisation chart."""
    apply_style()

    allocation, months, available = calculate_monthly_capacity(tasks, team, public_holidays, leave)
    if not months:
        print("  No monthly data. Check: tasks have positive Total Days and valid Start Date.")
        return

    person_list = list(team.keys())
    n_persons = len(person_list)
    n_months = len(months)

    fig, ax = plt.subplots(figsize=(max(14, n_months * 2.2), 8), facecolor=STYLE["bg_color"])

    x = np.arange(n_months)
    bar_width = 0.7 / n_persons

    # Grouped bars per person (per-person over-capacity colouring)
    for pidx, person in enumerate(person_list):
        allocated = [allocation[m][person] for m in months]
        offset = (pidx - (n_persons - 1) / 2) * bar_width

        colors = []
        for i, m in enumerate(months):
            person_alloc = allocation[m][person]
            person_avail = available[m].get(person, 0)
            if person_avail > 0 and person_alloc > person_avail:
                colors.append(STYLE["over_capacity_color"])
            else:
                colors.append(STYLE["under_capacity_colors"][pidx % len(STYLE["under_capacity_colors"])])

        ax.bar(x + offset, allocated, bar_width * 0.88,
               color=colors, alpha=0.85, edgecolor="white", linewidth=0.5,
               hatch=PERSON_HATCHES.get(pidx, ""), label=person)

    # Available capacity line
    total_avail_line = [sum(available[m].values()) for m in months]
    ax.plot(x, total_avail_line, color=STYLE["capacity_line_color"],
            linewidth=2, linestyle="--", marker="o", markersize=5,
            label="Available capacity", zorder=5)

    # Utilisation % labels
    for i, m in enumerate(months):
        total_alloc = sum(allocation[m].values())
        total_avail = sum(available[m].values())
        pct = (total_alloc / total_avail * 100) if total_avail > 0 else 0
        color = STYLE["over_capacity_color"] if pct > 100 else STYLE["text_secondary"]
        weight = "bold" if pct > 100 else "normal"
        y_label = max(total_alloc, total_avail_line[i]) + 1
        ax.text(i, y_label, f"{pct:.0f}%",
                ha="center", fontsize=STYLE["label_size"], color=color, fontweight=weight)

    ax.set_xticks(x)
    ax.set_xticklabels([m.strftime("%b %Y") for m in months], fontsize=STYLE["tick_size"])
    ax.legend(loc="upper right", fontsize=STYLE["small_size"], framealpha=0.9,
              edgecolor=STYLE["grid_color"], fancybox=True)

    # Annotate months with significant leave
    if leave:
        for i, m in enumerate(months):
            total_leave = 0
            for person in person_list:
                person_leave = leave.get(person, set())
                if person_leave:
                    _, num_days = monthrange(m.year, m.month)
                    for d in range(1, num_days + 1):
                        dt = datetime(m.year, m.month, d)
                        if dt.weekday() < 5 and dt in person_leave:
                            total_leave += 1
            if total_leave >= 3:  # Only annotate if 3+ leave days in month
                ax.text(i, -2, f"{total_leave}d leave",
                        ha="center", va="top", fontsize=6,
                        color=STYLE["leave_edge_color"], fontstyle="italic")

    max_val = max(max(sum(allocation[m].values()) for m in months), max(total_avail_line))
    ax.set_ylim(0, max_val * 1.2)
    style_axes(ax, title="Monthly Capacity Utilisation", ylabel="Working Days", show_grid_y=True)

    subtitle = f"{months[0].strftime('%b %Y')} \u2014 {months[-1].strftime('%b %Y')}"
    add_header_footer(fig, "Monthly Capacity Overview", subtitle)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fig.savefig(output_path, dpi=STYLE["dpi"], bbox_inches="tight", facecolor=STYLE["bg_color"])
    plt.close(fig)
    print(f"  Monthly chart saved: {output_path}")


# ── Chart: Strategic Roadmap ─────────────────────────────────────────────────

def render_roadmap(tasks, team, workstreams, output_path):
    """Render the strategic roadmap swim-lane view."""
    apply_style()

    ws_data = aggregate_workstreams(tasks, workstreams)
    if not ws_data:
        print("  No workstream data for roadmap.")
        return

    # Sort workstreams by priority then original order
    ws_order_items = list(workstreams.items())
    ws_order_items.sort(key=lambda item: (priority_sort_key(item[1]["priority"]),
                                          list(workstreams.keys()).index(item[0])))

    active_workstreams = [name for name, _ in ws_order_items if name in ws_data]
    n_workstreams = len(active_workstreams)

    fig, ax = plt.subplots(figsize=(STYLE["fig_width"], max(6, n_workstreams * 1.1 + 3)),
                           facecolor=STYLE["bg_color"])

    all_starts = [ws_data[s]["start"] for s in active_workstreams]
    all_ends = [ws_data[s]["end"] for s in active_workstreams]
    chart_start = min(all_starts) - timedelta(days=14)
    chart_end = max(all_ends) + timedelta(days=21)

    # Alternating row shading
    for i in range(n_workstreams):
        shade = STYLE["row_shade_even"] if i % 2 == 0 else STYLE["row_shade_odd"]
        ax.axhspan(i - 0.5, i + 0.5, color=shade, alpha=0.6, zorder=0)

    # Quarter boundaries
    q_starts = get_quarter_boundaries(chart_start, chart_end)
    for qs in q_starts:
        qs_num = mdates.date2num(qs)
        ax.axvline(qs_num, color=STYLE["grid_color"], linewidth=1.2, linestyle="-", alpha=0.5, zorder=1)
        ax.text(qs_num + 2, n_workstreams - 0.2, get_quarter_label(qs),
                fontsize=STYLE["tick_size"] + 1, color=STYLE["text_muted"],
                fontweight="bold", va="bottom", ha="left", zorder=6)

    # Draw workstream bars
    for idx, ws_name in enumerate(reversed(active_workstreams)):
        y = idx
        data = ws_data[ws_name]
        ws_info = workstreams[ws_name]
        color = ws_info["color"]
        ws_priority = ws_info["priority"]
        ws_pstyle = PRIORITY_STYLES.get(ws_priority, PRIORITY_STYLES["P2"])

        # Density-based segments (weekly)
        seg_start = data["start"]
        while seg_start <= data["end"]:
            seg_end = min(seg_start + timedelta(days=6), data["end"])

            active_count = sum(1 for d in data["day_counts"]
                               if seg_start <= d <= seg_end)
            max_possible = 5
            density_alpha = 0.25 + 0.65 * min(active_count / max(max_possible, 1), 1.0)

            s_num = mdates.date2num(seg_start)
            e_num = mdates.date2num(seg_end)
            width = max(e_num - s_num, 1)

            draw_rounded_bar(ax, s_num, y, width, 0.55, color,
                             alpha=density_alpha, edgecolor="none", linewidth=0, zorder=2)

            seg_start = seg_end + timedelta(days=1)

        # Full bar outline — linewidth varies by priority
        full_start = mdates.date2num(data["start"])
        full_end = mdates.date2num(data["end"])
        full_width = max(full_end - full_start, 1)
        draw_rounded_bar(ax, full_start, y, full_width, 0.55, "none",
                         alpha=1.0, edgecolor=color,
                         linewidth=ws_pstyle["linewidth"], zorder=3)

        # Task start markers (diamonds)
        for task_start, task_name in data["task_starts"]:
            ax.plot(mdates.date2num(task_start), y, marker="D",
                    markersize=5, color=color, markeredgecolor="white",
                    markeredgewidth=0.8, zorder=5)

        # Blocked warning markers
        for blocked_task in data.get("blocked_tasks", []):
            bt_start_num = mdates.date2num(blocked_task["start_date"])
            ax.plot(bt_start_num, y + 0.25, marker="$!$",
                    markersize=8, color=STYLE["late_color"],
                    zorder=6)

        # Task count label
        ax.text(full_end + 3, y, f"{data['task_count']} task{'s' if data['task_count'] != 1 else ''}",
                va="center", fontsize=STYLE["small_size"] + 0.5,
                color=STYLE["text_muted"], style="italic")

    # Y-axis: workstream names with priority badge
    y_labels = []
    for ws_name in reversed(active_workstreams):
        ws_priority = workstreams[ws_name]["priority"]
        y_labels.append(f"{ws_priority}  {ws_name}")

    ax.set_yticks(range(n_workstreams))
    ax.set_yticklabels(y_labels, fontsize=STYLE["label_size"])
    for tick_label, ws_name in zip(ax.get_yticklabels(), reversed(active_workstreams)):
        tick_label.set_color(workstreams[ws_name]["color"])
        tick_label.set_fontweight("bold")

    # X-axis: months
    ax.xaxis.set_major_locator(mdates.MonthLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.set_xlim(mdates.date2num(chart_start), mdates.date2num(chart_end))
    plt.setp(ax.xaxis.get_majorticklabels(), fontsize=STYLE["tick_size"])

    draw_today_line(ax, chart_start, chart_end, n_workstreams - 0.3)
    style_axes(ax, title="Strategic Roadmap")

    # Legend
    diamond_handle = plt.Line2D([0], [0], marker="D", color="w", markerfacecolor="#666666",
                                markeredgecolor="white", markersize=6, label="Task start",
                                linestyle="None")
    shade_handle = mpatches.Patch(facecolor="#888888", alpha=0.5, edgecolor="#888888",
                                  label="Activity density")
    blocked_handle = plt.Line2D([0], [0], marker="$!$", color="w",
                                markerfacecolor=STYLE["late_color"], markersize=8,
                                label="Blocked task", linestyle="None")
    ax.legend(handles=[diamond_handle, shade_handle, blocked_handle], loc="upper right",
              fontsize=STYLE["small_size"], framealpha=0.9,
              edgecolor=STYLE["grid_color"], fancybox=True)

    subtitle = f"{chart_start.strftime('%b %Y')} \u2014 {chart_end.strftime('%b %Y')}"
    add_header_footer(fig, "Strategic Roadmap", subtitle)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fig.savefig(output_path, dpi=STYLE["dpi"], bbox_inches="tight", facecolor=STYLE["bg_color"])
    plt.close(fig)
    print(f"  Roadmap saved: {output_path}")


# ── Executive Summary ────────────────────────────────────────────────────────

def print_summary(tasks, team, workstreams, allocation, weeks, available=None,
                  public_holidays=None, leave=None, leave_entries=None):
    """Print executive summary statistics to console."""
    total_tasks = len(tasks)
    active = len([t for t in tasks if t["status"] in ("Planned", "In Progress")])
    complete = len([t for t in tasks if t["status"] == "Complete"])
    on_hold = len([t for t in tasks if t["status"] == "On Hold"])

    over_weeks = 0
    over_capacity_detail = {}  # {person: [(week, alloc, avail), ...]}
    person_totals = {name: 0.0 for name in team}
    person_available_totals = {name: 0.0 for name in team}
    for w in weeks:
        week_team_avail = sum(available[w][p] for p in team) if available and w in available else sum(team.values())
        week_total = sum(allocation[w].values())
        if week_total > week_team_avail:
            over_weeks += 1
        for person in team:
            alloc = allocation[w].get(person, 0)
            avail = available[w][person] if available and w in available else team[person]
            person_totals[person] += alloc
            person_available_totals[person] += avail
            if avail > 0 and alloc > avail:
                over_capacity_detail.setdefault(person, []).append((w, alloc, avail))

    total_available = sum(person_available_totals.values())
    total_allocated = sum(person_totals.values())
    overall_util = (total_allocated / total_available * 100) if total_available > 0 else 0

    busiest = max(person_totals, key=person_totals.get) if person_totals else "N/A"

    print()
    print("=" * 60)
    print("  EXECUTIVE SUMMARY")
    print("=" * 60)
    print(f"  Tasks:         {total_tasks} total ({active} active, "
          f"{complete} complete, {on_hold} on hold)")
    print(f"  Timeline:      {len(weeks)} weeks")
    if public_holidays:
        if weeks:
            timeline_start = weeks[0]
            timeline_end = weeks[-1] + timedelta(days=4)
            hols_in_range = sum(1 for h in public_holidays if timeline_start <= h <= timeline_end)
            if hols_in_range:
                print(f"  Public holidays: {hols_in_range} in timeline period")
    print(f"  Utilisation:   {overall_util:.0f}% overall")
    for person in team:
        p_avail = person_available_totals[person]
        p_util = (person_totals[person] / p_avail * 100) if p_avail > 0 else 0
        print(f"    {person}: {person_totals[person]:.1f} / {p_avail:.0f} days ({p_util:.0f}%)")
    print(f"  Busiest:       {busiest} ({person_totals.get(busiest, 0):.1f} days allocated)")
    print(f"  Over-capacity: {over_weeks} of {len(weeks)} weeks")
    if over_weeks > 0:
        for person, entries in over_capacity_detail.items():
            detail_strs = [f"w/c {w.strftime('%d %b')} ({alloc:.1f}d vs {avail:.1f}d)"
                           for w, alloc, avail in entries[:3]]
            suffix = f" ... +{len(entries) - 3} more" if len(entries) > 3 else ""
            print(f"    {person}: {', '.join(detail_strs)}{suffix}")

    # Leave summary
    if leave_entries:
        print()
        print("  Leave:")
        # Group by person
        by_person = {}
        for entry in leave_entries:
            by_person.setdefault(entry["person"], []).append(entry)
        for person, entries in by_person.items():
            parts = []
            for e in entries:
                parts.append(f"{e['days']}d {e['type']} ({e['start'].strftime('%d %b')} - {e['end'].strftime('%d %b')})")
            print(f"    {person}: {', '.join(parts)}")

    # Priority breakdown
    print()
    print("  By priority:")
    for p in PRIORITY_VALUES:
        p_tasks = [t for t in tasks if t["priority"] == p]
        if p_tasks:
            p_days = sum(t["total_days"] for t in p_tasks)
            print(f"    {p}: {len(p_tasks)} task{'s' if len(p_tasks) != 1 else ''} ({p_days:.4g} days)")

    # Deadline warnings
    deadline_tasks = [t for t in tasks if t.get("deadline") and t.get("end_date")]
    at_risk = [t for t in deadline_tasks if t["end_date"] > t["deadline"]]
    if at_risk:
        print()
        print(f"  Deadlines at risk: {len(at_risk)}")
        for t in at_risk:
            person_leave = leave.get(t["assigned_to"], set()) if leave else None
            overshoot_days = count_working_days(
                t["deadline"] + timedelta(days=1), t["end_date"],
                public_holidays, person_leave)
            print(f"    WARNING: '{t['task']}' ends {overshoot_days} wd after deadline "
                  f"(deadline: {t['deadline'].strftime('%d %b')})")

    # Low confidence tasks
    low_conf = [t for t in tasks if t.get("confidence") == "Low"]
    if low_conf:
        print()
        print(f"  Low confidence estimates: {len(low_conf)}")
        for t in low_conf:
            print(f"    {t['task']} ({t['total_days']:.4g} wd)")

    # Estimation drift total
    drift_tasks = [t for t in tasks if t["original_days"] != t["total_days"] and t["original_days"] > 0]
    if drift_tasks:
        total_original = sum(t["original_days"] for t in drift_tasks)
        total_current = sum(t["total_days"] for t in drift_tasks)
        total_drift = total_current - total_original
        drift_pct = (total_drift / total_original) * 100 if total_original > 0 else 0
        sign = "+" if total_drift > 0 else ""
        direction = "increase" if total_drift > 0 else "decrease"
        print()
        print(f"  Estimation drift: {sign}{total_drift:.4g} days ({sign}{drift_pct:.0f}%) "
              f"across {len(drift_tasks)} task{'s' if len(drift_tasks) != 1 else ''} — scope {direction}")
        for t in drift_tasks:
            pct = ((t["total_days"] - t["original_days"]) / t["original_days"]) * 100
            s = "+" if pct > 0 else ""
            print(f"    {t['task']}: {t['original_days']:.4g}d -> {t['total_days']:.4g}d ({s}{pct:.0f}%)")

    # Concurrent task awareness
    if weeks and allocation:
        for person in team:
            for w in weeks:
                concurrent = sum(1 for t in tasks
                                 if t["assigned_to"] == person
                                 and t.get("working_days")
                                 and any(norm_date(wd) >= w and norm_date(wd) < w + timedelta(days=5)
                                         for wd in t["working_days"]))
                if concurrent >= 3:
                    print(f"  NOTE: {person} has {concurrent} concurrent tasks in w/c {w.strftime('%d %b')}")
                    break  # One note per person is enough

    print("=" * 60)
    print()


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Capacity Planning Tool \u2014 generate Gantt, capacity, and roadmap charts from Excel data"
    )
    parser.add_argument(
        "--template", action="store_true",
        help="Generate a blank Excel template with example data"
    )
    parser.add_argument(
        "--input", default=DEFAULT_INPUT,
        help="Path to Excel input file (default: capacity_data.xlsx)"
    )
    parser.add_argument(
        "--output", default=None,
        help="[Deprecated] Use --outdir instead. If provided, treated as output directory."
    )
    parser.add_argument(
        "--outdir", default=None,
        help="Output directory for all charts (default: output/)"
    )
    parser.add_argument(
        "--charts", default=["all"], nargs="+",
        choices=["all", "gantt", "weekly", "monthly", "roadmap"],
        help="Which charts to generate (default: all). Can specify multiple: --charts gantt weekly monthly"
    )
    parser.add_argument(
        "--from", dest="date_from", default=None,
        help="Only include tasks overlapping with this start date (YYYY-MM-DD)"
    )
    parser.add_argument(
        "--to", dest="date_to", default=None,
        help="Only include tasks overlapping with this end date (YYYY-MM-DD)"
    )
    args = parser.parse_args()

    if args.template:
        generate_template(args.input)
        return

    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}")
        print("Run with --template first to create a template.")
        sys.exit(1)

    # Determine output directory
    if args.outdir:
        out_dir = args.outdir
    elif args.output:
        # Deprecated --output: treat as directory
        out_dir = args.output
        if os.path.splitext(args.output)[1]:  # has file extension — use parent dir
            out_dir = os.path.dirname(args.output) or "output"
        print(f"  NOTE: --output is deprecated. Use --outdir instead.")
    else:
        out_dir = os.path.join(_DIR, "output")

    gantt_path = os.path.join(out_dir, "capacity_gantt.png")
    weekly_path = os.path.join(out_dir, "capacity_weekly.png")
    monthly_path = os.path.join(out_dir, "capacity_monthly.png")
    roadmap_path = os.path.join(out_dir, "roadmap.png")

    # Parse date window
    date_from = None
    date_to = None
    if args.date_from:
        try:
            date_from = datetime.strptime(args.date_from, "%Y-%m-%d")
        except ValueError:
            print(f"  ERROR: Invalid --from date '{args.date_from}'. Use YYYY-MM-DD format.")
            sys.exit(1)
    if args.date_to:
        try:
            date_to = datetime.strptime(args.date_to, "%Y-%m-%d")
        except ValueError:
            print(f"  ERROR: Invalid --to date '{args.date_to}'. Use YYYY-MM-DD format.")
            sys.exit(1)

    # Load
    print(f"Loading data from: {args.input}")
    try:
        mtime = datetime.fromtimestamp(os.path.getmtime(args.input))
        STYLE["_data_mtime"] = mtime.strftime("%d %b %Y %H:%M")
    except OSError:
        pass
    team, workstreams, tasks, public_holidays, leave, leave_entries = load_data(args.input)
    print(f"  Team: {', '.join(f'{n} ({d}d/wk)' for n, d in team.items())}")
    print(f"  Workstreams: {len(workstreams)}")
    print(f"  Tasks: {len(tasks)}")

    # Validate
    errors, warnings = validate_data(team, workstreams, tasks, public_holidays, leave)
    for w in warnings:
        print(f"  WARNING: {w}")
    if errors:
        for e in errors:
            print(f"  ERROR: {e}")
        sys.exit(1)

    # Calculate
    tasks = calculate_schedule(tasks, public_holidays, leave)

    # Post-schedule integrity check
    for t in tasks:
        if t.get("end_date") and t["end_date"] < t["start_date"]:
            print(f"  WARNING: Task '{t['task']}' has end_date before start_date after scheduling. "
                  f"Check your data.")

    # Apply date window filter
    if date_from or date_to:
        original_count = len(tasks)
        filtered = []
        for t in tasks:
            t_start = t["start_date"]
            t_end = t["end_date"]
            if date_from and t_end < date_from:
                continue  # task ends before window
            if date_to and t_start > date_to:
                continue  # task starts after window
            filtered.append(t)
        tasks = filtered
        if len(tasks) < original_count:
            window_desc = ""
            if date_from:
                window_desc += f"from {date_from.strftime('%d %b %Y')}"
            if date_to:
                window_desc += f"{' ' if window_desc else ''}to {date_to.strftime('%d %b %Y')}"
            print(f"  Date filter ({window_desc}): {len(tasks)} of {original_count} tasks")
        if not tasks:
            print(f"  WARNING: 0 tasks overlap with date window. No charts will be generated.")

    allocation, weeks, available = calculate_capacity(tasks, team, public_holidays, leave)
    print(f"  Weeks covered: {len(weeks)}")

    # Summary + schedule suggestions (capture output for summary.txt)
    summary_capture = io.StringIO()
    _orig_stdout = sys.stdout
    sys.stdout = _TeeWriter(_orig_stdout, summary_capture)
    print_summary(tasks, team, workstreams, allocation, weeks, available,
                  public_holidays, leave, leave_entries)
    print_schedule_suggestions(tasks, team, allocation, weeks, available,
                               public_holidays, leave)
    sys.stdout = _orig_stdout
    summary_text = summary_capture.getvalue()

    # Determine which charts
    charts = args.charts
    gen_all = "all" in charts
    output_files = []

    # Render
    if gen_all or "gantt" in charts:
        render_gantt(tasks, team, workstreams, weeks, gantt_path,
                     public_holidays, leave)
        output_files.append(gantt_path)

    if gen_all or "weekly" in charts:
        render_weekly(tasks, team, workstreams, allocation, weeks, available,
                      weekly_path, public_holidays, leave)
        output_files.append(weekly_path)

    if gen_all or "monthly" in charts:
        render_monthly_capacity(tasks, team, workstreams, monthly_path,
                                public_holidays, leave)
        output_files.append(monthly_path)

    if gen_all or "roadmap" in charts:
        render_roadmap(tasks, team, workstreams, roadmap_path)
        output_files.append(roadmap_path)

    # Write summary.txt
    summary_path = os.path.join(out_dir, "summary.txt")
    with open(summary_path, "w", encoding="utf-8") as sf:
        sf.write(summary_text)
    output_files.append(summary_path)

    # Output summary
    if output_files:
        print()
        print("  Output:")
        for f in output_files:
            print(f"    {os.path.abspath(f)}")
    print("\nDone.")


if __name__ == "__main__":
    main()
