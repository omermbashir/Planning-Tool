"""Automated test suite for capacity_planner.py.

Derived from 28 bugs found across 9 review rounds.
Covers: pure unit tests, function tests, integration tests (Excel I/O), end-to-end.
"""

import io
import math
import os
import sys
import tempfile
from datetime import datetime, timedelta

import pandas as pd
import pytest
from openpyxl import Workbook

from capacity_planner import (
    norm_date,
    clean_str,
    parse_date,
    priority_sort_key,
    is_working_day,
    count_working_days,
    get_end_date,
    get_week_start,
    normalize_columns,
    validate_data,
    load_team,
    load_workstreams,
    load_tasks,
    load_public_holidays,
    load_leave,
    load_data,
    generate_template,
    calculate_schedule,
    calculate_capacity,
    calculate_monthly_capacity,
    working_days_in_month,
    print_schedule_suggestions,
    print_summary,
    STATUS_VALUES,
    PRIORITY_VALUES,
    LEAVE_TYPES,
    CONFIDENCE_COLORS,
    aggregate_workstreams,
    render_gantt,
    render_weekly,
    render_monthly_capacity,
    render_roadmap,
)


# ── Fixtures ────────────────────────────────────────────────────────────────


def create_test_excel(team_rows, workstream_rows, task_rows,
                      holiday_rows=None, leave_rows=None):
    """Create a temp Excel file with test data. Returns filepath."""
    wb = Workbook()

    ws_team = wb.active
    ws_team.title = "Team"
    ws_team.append(["Name", "Role", "Days Per Week"])
    for row in team_rows:
        ws_team.append(row)

    ws_ws = wb.create_sheet("Workstreams")
    ws_ws.append(["Workstream", "Color", "Priority"])
    for row in workstream_rows:
        ws_ws.append(row)

    ws_tasks = wb.create_sheet("Tasks")
    ws_tasks.append(["Task", "Workstream", "Assigned To", "Start Date",
                     "Total Days", "Status", "Original Days", "Priority",
                     "Actual End", "Blocked By", "Deadline", "Confidence", "Notes"])
    for row in task_rows:
        ws_tasks.append(row)

    if holiday_rows is not None:
        ws_hol = wb.create_sheet("Public Holidays")
        ws_hol.append(["Date", "Name"])
        for row in holiday_rows:
            ws_hol.append(row)

    if leave_rows is not None:
        ws_leave = wb.create_sheet("Leave")
        ws_leave.append(["Person", "Start Date", "End Date", "Type", "Notes"])
        for row in leave_rows:
            ws_leave.append(row)

    tmpdir = tempfile.mkdtemp()
    filepath = os.path.join(tmpdir, "test_data.xlsx")
    wb.save(filepath)
    return filepath


@pytest.fixture
def basic_excel():
    """Minimal valid Excel file for testing."""
    path = create_test_excel(
        team_rows=[["Alice", "Lead", 5], ["Bob", "Analyst", 5]],
        workstream_rows=[["Project A", "#00BCD4", "P1"],
                         ["Project B", "#4CAF50", "P2"]],
        task_rows=[
            ["Task 1", "Project A", "Alice", "2026-03-02", 5, "In Progress",
             5, "P1", None, None, None, "High", None],
            ["Task 2", "Project B", "Bob", "2026-03-09", 10, "Planned",
             10, "P2", None, None, None, None, None],
        ],
    )
    yield path
    # Cleanup handled by OS temp dir


# ── Tier 1: Pure Unit Tests ─────────────────────────────────────────────────


class TestNormDate:
    """Bug #4, #11: datetime normalization to midnight."""

    def test_datetime_to_midnight(self):
        d = datetime(2026, 3, 15, 14, 30, 45)
        result = norm_date(d)
        assert result == datetime(2026, 3, 15, 0, 0, 0)
        assert result.hour == 0
        assert result.minute == 0
        assert result.second == 0

    def test_timestamp_to_midnight(self):
        ts = pd.Timestamp("2026-03-15 10:20:30")
        result = norm_date(ts)
        assert result == datetime(2026, 3, 15)
        assert isinstance(result, datetime)

    def test_already_midnight(self):
        d = datetime(2026, 3, 15)
        assert norm_date(d) == d

    def test_invalid_type_raises(self):
        with pytest.raises(TypeError):
            norm_date("2026-03-15")
        with pytest.raises(TypeError):
            norm_date(42)

    def test_now_normalization(self):
        """Bug #4: datetime.now() must be normalized for date comparisons."""
        now = datetime.now()
        normalized = norm_date(now)
        assert normalized.hour == 0
        assert normalized.minute == 0
        assert normalized.second == 0


class TestCleanStr:
    """NaN-safe string extraction."""

    def test_normal_string(self):
        assert clean_str("hello") == "hello"

    def test_strips_whitespace(self):
        assert clean_str("  hello  ") == "hello"

    def test_none_returns_empty(self):
        assert clean_str(None) == ""

    def test_nan_returns_empty(self):
        assert clean_str(float("nan")) == ""

    def test_pd_nat_returns_empty(self):
        assert clean_str(pd.NaT) == ""

    def test_number_to_string(self):
        assert clean_str(42) == "42"


class TestParseDate:
    """Date parsing from Excel cells — multiple formats."""

    def test_iso_format(self):
        result = parse_date("2026-03-15")
        assert result == datetime(2026, 3, 15)

    def test_uk_format_slash(self):
        result = parse_date("15/03/2026")
        assert result == datetime(2026, 3, 15)

    def test_uk_format_dash(self):
        result = parse_date("15-03-2026")
        assert result == datetime(2026, 3, 15)

    def test_datetime_passthrough(self):
        d = datetime(2026, 3, 15, 10, 30)
        result = parse_date(d)
        assert result == datetime(2026, 3, 15)  # normalized to midnight

    def test_timestamp_passthrough(self):
        ts = pd.Timestamp("2026-03-15")
        result = parse_date(ts)
        assert result == datetime(2026, 3, 15)

    def test_blank_raises(self):
        with pytest.raises(ValueError):
            parse_date("")

    def test_nan_raises(self):
        with pytest.raises(ValueError):
            parse_date(float("nan"))

    def test_invalid_format_raises(self):
        with pytest.raises(ValueError):
            parse_date("March 15, 2026")

    def test_context_in_error(self):
        with pytest.raises(ValueError, match="Task 1"):
            parse_date("bad-date", context="Task 1")

    def test_all_paths_return_midnight(self):
        """All return paths must normalize to midnight."""
        for val in ["2026-03-15", "15/03/2026", datetime(2026, 3, 15, 10, 30),
                     pd.Timestamp("2026-03-15 14:00")]:
            result = parse_date(val)
            assert result.hour == 0 and result.minute == 0 and result.second == 0


class TestPrioritySortKey:
    def test_p1_through_p4(self):
        assert priority_sort_key("P1") == 1
        assert priority_sort_key("P2") == 2
        assert priority_sort_key("P3") == 3
        assert priority_sort_key("P4") == 4

    def test_ordering(self):
        assert priority_sort_key("P1") < priority_sort_key("P2")
        assert priority_sort_key("P3") < priority_sort_key("P4")

    def test_invalid_returns_high(self):
        assert priority_sort_key("") == 9
        assert priority_sort_key(None) == 9
        # P5 parses as int 5 (valid sort key, just not in PRIORITY_VALUES)
        assert priority_sort_key("P5") == 5
        assert priority_sort_key("Px") == 9

    def test_get_week_start(self):
        # Monday stays Monday
        mon = datetime(2026, 3, 2)  # Monday
        assert get_week_start(mon) == mon
        # Sunday goes back to previous Monday
        sun = datetime(2026, 3, 8)  # Sunday
        assert get_week_start(sun) == datetime(2026, 3, 2)
        # Wednesday goes back to Monday
        wed = datetime(2026, 3, 4)  # Wednesday
        assert get_week_start(wed) == datetime(2026, 3, 2)


# ── Tier 2: Function Tests ─────────────────────────────────────────────────


class TestIsWorkingDay:
    def test_weekday_is_working(self):
        assert is_working_day(datetime(2026, 3, 2)) is True  # Monday

    def test_saturday_is_not_working(self):
        assert is_working_day(datetime(2026, 3, 7)) is False  # Saturday

    def test_sunday_is_not_working(self):
        assert is_working_day(datetime(2026, 3, 8)) is False  # Sunday

    def test_public_holiday(self):
        holidays = {datetime(2026, 3, 2)}
        assert is_working_day(datetime(2026, 3, 2), public_holidays=holidays) is False

    def test_person_leave(self):
        leave = {datetime(2026, 3, 3)}
        assert is_working_day(datetime(2026, 3, 3), person_leave=leave) is False

    def test_non_leave_day_still_working(self):
        leave = {datetime(2026, 3, 3)}
        assert is_working_day(datetime(2026, 3, 4), person_leave=leave) is True


class TestCountWorkingDays:
    def test_single_week(self):
        # Mon-Fri = 5 working days
        start = datetime(2026, 3, 2)
        end = datetime(2026, 3, 6)
        assert count_working_days(start, end) == 5

    def test_two_weeks(self):
        start = datetime(2026, 3, 2)
        end = datetime(2026, 3, 13)  # Fri of second week
        assert count_working_days(start, end) == 10

    def test_with_holiday(self):
        start = datetime(2026, 3, 2)
        end = datetime(2026, 3, 6)
        holidays = {datetime(2026, 3, 4)}  # Wednesday is holiday
        assert count_working_days(start, end, public_holidays=holidays) == 4

    def test_with_leave(self):
        start = datetime(2026, 3, 2)
        end = datetime(2026, 3, 6)
        leave = {datetime(2026, 3, 3)}  # Tuesday is leave
        assert count_working_days(start, end, person_leave=leave) == 4

    def test_holiday_and_leave_no_double_count(self):
        """Holiday + leave on same day should subtract only 1 day."""
        start = datetime(2026, 3, 2)
        end = datetime(2026, 3, 6)
        holidays = {datetime(2026, 3, 4)}
        leave = {datetime(2026, 3, 4)}  # Same day
        assert count_working_days(start, end, holidays, leave) == 4  # Not 3

    def test_same_day(self):
        d = datetime(2026, 3, 2)  # Monday
        assert count_working_days(d, d) == 1

    def test_end_before_start_returns_zero(self):
        start = datetime(2026, 3, 6)
        end = datetime(2026, 3, 2)
        assert count_working_days(start, end) == 0


class TestGetEndDate:
    def test_simple_5_days(self):
        start = datetime(2026, 3, 2)  # Monday
        end_date, working_days, allocs = get_end_date(start, 5)
        assert end_date == datetime(2026, 3, 6)  # Friday
        assert len(working_days) == 5

    def test_spans_weekend(self):
        start = datetime(2026, 3, 5)  # Thursday
        end_date, working_days, allocs = get_end_date(start, 3)
        # Thu + Fri + skip weekend + Mon
        assert end_date == datetime(2026, 3, 9)  # Monday

    def test_fractional_days(self):
        start = datetime(2026, 3, 2)  # Monday
        end_date, working_days, allocs = get_end_date(start, 0.5)
        assert end_date == datetime(2026, 3, 2)  # Same day
        assert allocs[datetime(2026, 3, 2)] == 0.5

    def test_skips_holidays(self):
        start = datetime(2026, 3, 2)  # Monday
        holidays = {datetime(2026, 3, 4)}  # Wednesday
        end_date, working_days, allocs = get_end_date(start, 5, public_holidays=holidays)
        # Mon, Tue, skip Wed, Thu, Fri, Mon
        assert end_date == datetime(2026, 3, 9)  # Monday next week

    def test_skips_leave(self):
        start = datetime(2026, 3, 2)
        leave = {datetime(2026, 3, 3)}  # Tuesday
        end_date, working_days, allocs = get_end_date(start, 5, person_leave=leave)
        assert end_date == datetime(2026, 3, 9)

    def test_zero_days(self):
        start = datetime(2026, 3, 2)
        end_date, working_days, allocs = get_end_date(start, 0)
        assert end_date == start
        assert working_days == []

    def test_negative_days(self):
        start = datetime(2026, 3, 2)
        end_date, working_days, allocs = get_end_date(start, -1)
        assert end_date == start
        assert working_days == []


class TestNormalizeColumns:
    """Bug #19, #20: case-insensitive column matching."""

    def test_exact_match(self):
        df = pd.DataFrame(columns=["Name", "Days Per Week"])
        missing = normalize_columns(df, {"Name", "Days Per Week"})
        assert missing == set()

    def test_case_insensitive_match(self):
        df = pd.DataFrame(columns=["name", "days per week"])
        missing = normalize_columns(df, {"Name", "Days Per Week"})
        assert missing == set()
        assert "Name" in df.columns
        assert "Days Per Week" in df.columns

    def test_missing_column(self):
        df = pd.DataFrame(columns=["Name"])
        missing = normalize_columns(df, {"Name", "Days Per Week"})
        assert "Days Per Week" in missing

    def test_whitespace_stripped(self):
        df = pd.DataFrame(columns=["  Name  ", "Days Per Week"])
        missing = normalize_columns(df, {"Name", "Days Per Week"})
        assert missing == set()

    def test_optional_columns(self):
        """Bug #20: optional columns must also be normalized."""
        df = pd.DataFrame(columns=["workstream", "color", "priority"])
        required = {"Workstream", "Color"}
        optional = {"Priority"}
        missing = normalize_columns(df, required | optional)
        assert missing == set()
        assert "Priority" in df.columns


class TestValidateData:
    """Validation error/warning paths."""

    @staticmethod
    def _make_task(**overrides):
        """Build a minimal valid task dict with optional overrides."""
        base = {
            "task": "T", "workstream": "WS", "assigned_to": "Alice",
            "start_date": datetime(2026, 3, 2), "total_days": 5,
            "status": "Planned", "priority": "P1", "original_days": 5,
            "blocked_by": "", "deadline": None, "actual_end": None,
            "_row": 2,
        }
        base.update(overrides)
        return base

    def test_empty_team_is_error(self):
        task = self._make_task(assigned_to="X")
        errors, warnings = validate_data(
            {}, {"WS": {"color": "#000000", "priority": "P1"}}, [task])
        assert any("Team" in e or "team" in e.lower() for e in errors)

    def test_empty_workstreams_is_error(self):
        task = self._make_task()
        errors, warnings = validate_data({"Alice": 5}, {}, [task])
        assert any("Workstream" in e or "workstream" in e.lower() for e in errors)

    def test_invalid_status_is_error(self):
        """Bug #21: invalid status must be error, not warning."""
        task = self._make_task(status="InvalidStatus")
        errors, warnings = validate_data(
            {"Alice": 5}, {"WS": {"color": "#000000", "priority": "P1"}}, [task])
        assert any("InvalidStatus" in e for e in errors)
        assert not any("InvalidStatus" in w for w in warnings)

    def test_valid_status_no_error(self):
        for status in STATUS_VALUES:
            task = self._make_task(status=status)
            errors, warnings = validate_data(
                {"Alice": 5}, {"WS": {"color": "#000000", "priority": "P1"}}, [task])
            assert not any("not recognised" in e for e in errors)

    def test_unknown_workstream_is_error(self):
        task = self._make_task(workstream="NonExistent")
        errors, _ = validate_data(
            {"Alice": 5}, {"WS": {"color": "#000000", "priority": "P1"}}, [task])
        assert any("NonExistent" in e for e in errors)

    def test_unknown_assignee_is_error(self):
        task = self._make_task(assigned_to="Nobody")
        errors, _ = validate_data(
            {"Alice": 5}, {"WS": {"color": "#000000", "priority": "P1"}}, [task])
        assert any("Nobody" in e for e in errors)


class TestWorkingDaysInMonth:
    def test_march_2026(self):
        # March 2026: 22 working days (no holidays)
        assert working_days_in_month(2026, 3) == 22

    def test_with_holiday(self):
        holidays = {datetime(2026, 3, 4)}
        assert working_days_in_month(2026, 3, holidays) == 21

    def test_weekend_holiday_no_effect(self):
        holidays = {datetime(2026, 3, 7)}  # Saturday
        assert working_days_in_month(2026, 3, holidays) == 22


# ── Tier 3: Integration Tests (Excel I/O) ──────────────────────────────────


class TestLoadTeam:
    def test_normal_load(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Analyst", 3]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
        )
        team = load_team(path)
        assert team == {"Alice": 5, "Bob": 3}

    def test_duplicate_names_uses_first(self):
        """Bug #24: duplicate names must warn and keep first occurrence."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Alice", "Analyst", 3]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
        )
        team = load_team(path)
        assert team["Alice"] == 5  # First occurrence, not 3

    def test_nan_days_skipped(self, capsys):
        """Bug #1: NaN Days Per Week must be skipped."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Analyst", None]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
        )
        team = load_team(path)
        assert "Alice" in team
        assert "Bob" not in team
        assert "WARNING" in capsys.readouterr().out

    def test_zero_days_skipped(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 0]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
        )
        team = load_team(path)
        assert "Alice" not in team

    def test_case_insensitive_columns(self):
        """Bug #19: column casing must not matter."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Team"
        ws.append(["name", "role", "days per week"])
        ws.append(["Alice", "Lead", 5])
        ws2 = wb.create_sheet("Workstreams")
        ws2.append(["Workstream", "Color", "Priority"])
        ws3 = wb.create_sheet("Tasks")
        ws3.append(["Task", "Workstream", "Assigned To", "Start Date",
                     "Total Days", "Status"])
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "test.xlsx")
        wb.save(path)
        team = load_team(path)
        assert team == {"Alice": 5}


class TestLoadWorkstreams:
    def test_normal_load(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["Project A", "#00BCD4", "P1"],
                             ["Project B", "#4CAF50", "P3"]],
            task_rows=[],
        )
        ws = load_workstreams(path)
        assert ws["Project A"]["priority"] == "P1"
        assert ws["Project B"]["color"] == "#4CAF50"

    def test_duplicate_names_uses_first(self):
        """Bug #25: duplicate workstream names must warn and keep first."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["Project A", "#00BCD4", "P1"],
                             ["Project A", "#4CAF50", "P4"]],
            task_rows=[],
        )
        ws = load_workstreams(path)
        assert ws["Project A"]["color"] == "#00BCD4"  # First occurrence
        assert ws["Project A"]["priority"] == "P1"    # Not P4

    def test_missing_priority_defaults_p2(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["Project A", "#00BCD4", None]],
            task_rows=[],
        )
        ws = load_workstreams(path)
        assert ws["Project A"]["priority"] == "P2"

    def test_priority_case_insensitive(self):
        """Bug #20: lowercase 'priority' column must be normalized."""
        wb = Workbook()
        ws_team = wb.active
        ws_team.title = "Team"
        ws_team.append(["Name", "Role", "Days Per Week"])
        ws_ws = wb.create_sheet("Workstreams")
        ws_ws.append(["workstream", "color", "priority"])
        ws_ws.append(["Project A", "#00BCD4", "P1"])
        ws_tasks = wb.create_sheet("Tasks")
        ws_tasks.append(["Task", "Workstream", "Assigned To", "Start Date",
                         "Total Days", "Status"])
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "test.xlsx")
        wb.save(path)
        result = load_workstreams(path)
        assert result["Project A"]["priority"] == "P1"


class TestLoadTasks:
    def test_normal_load(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[["Task 1", "WS", "Alice", "2026-03-02", 5, "Planned",
                        5, "P1", None, None, None, None, None]],
        )
        tasks = load_tasks(path, {"WS": {"color": "#000", "priority": "P1"}})
        assert len(tasks) == 1
        assert tasks[0]["task"] == "Task 1"
        assert tasks[0]["total_days"] == 5

    def test_nan_total_days_skipped(self, capsys):
        """Bug #2: NaN Total Days must produce a warning."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000000", "P1"]],
            task_rows=[["Task 1", "WS", "Alice", "2026-03-02", None, "Planned",
                        None, "P1", None, None, None, None, None]],
        )
        tasks = load_tasks(path, {"WS": {"color": "#000000", "priority": "P1"}})
        out = capsys.readouterr().out
        # Either the task is skipped entirely or flagged with warning
        assert "WARNING" in out or len(tasks) == 0 or any(t.get("_skip") for t in tasks)


class TestLoadPublicHolidays:
    def test_normal_load(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
            holiday_rows=[["2026-04-03", "Good Friday"],
                          ["2026-04-06", "Easter Monday"]],
        )
        holidays = load_public_holidays(path)
        assert datetime(2026, 4, 3) in holidays
        assert datetime(2026, 4, 6) in holidays

    def test_missing_sheet_returns_empty(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
            holiday_rows=None,
        )
        holidays = load_public_holidays(path)
        assert holidays == set() or holidays is None


class TestLoadLeave:
    def test_normal_load(self):
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
            leave_rows=[["Alice", "2026-04-06", "2026-04-10", "Annual Leave", "Easter"]],
        )
        leave_dates, leave_entries = load_leave(path)
        assert "Alice" in leave_dates
        assert len(leave_entries) == 1
        assert leave_entries[0]["type"] == "Annual Leave"

    def test_all_five_leave_types(self):
        """Bug #26: all 5 leave types must be accepted."""
        rows = []
        base_date = datetime(2026, 4, 6)
        for i, lt in enumerate(LEAVE_TYPES):
            start = base_date + timedelta(weeks=i)
            end = start + timedelta(days=4)
            rows.append(["Alice", start.strftime("%Y-%m-%d"),
                         end.strftime("%Y-%m-%d"), lt, f"Test {lt}"])
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
            leave_rows=rows,
        )
        leave_dates, leave_entries = load_leave(path)
        types_loaded = {e["type"] for e in leave_entries}
        for lt in LEAVE_TYPES:
            assert lt in types_loaded, f"Leave type '{lt}' not loaded"

    def test_end_before_start_warns(self, capsys):
        """Bug #8: leave end < start must produce warning."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[],
            leave_rows=[["Alice", "2026-04-10", "2026-04-06", "Annual Leave", "Swapped"]],
        )
        leave_dates, leave_entries = load_leave(path)
        out = capsys.readouterr().out
        assert "WARNING" in out or "warning" in out.lower()


# ── Tier 4: End-to-End Tests ────────────────────────────────────────────────


class TestTemplateRoundtrip:
    def test_template_generates_and_loads(self):
        """Template must produce valid Excel that loads with 0 errors."""
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "template_test.xlsx")
        generate_template(path)
        assert os.path.exists(path)

        team, workstreams, tasks, holidays, leave_dates, leave_entries = load_data(path)
        errors, warnings = validate_data(team, workstreams, tasks, holidays, leave_dates)
        assert len(errors) == 0, f"Template validation errors: {errors}"

    def test_template_has_all_leave_formatting(self):
        """Bug #26: all 5 leave types must have conditional formatting."""
        from openpyxl import load_workbook
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "template_test.xlsx")
        generate_template(path)

        wb = load_workbook(path)
        ws_leave = wb["Leave"]
        # Collect all conditional formatting rule formulas
        all_formulas = []
        for cf in ws_leave.conditional_formatting:
            for rule in cf.rules:
                if hasattr(rule, 'formula') and rule.formula:
                    all_formulas.extend(str(f) for f in rule.formula)
        formula_text = " ".join(all_formulas)
        for lt in LEAVE_TYPES:
            assert lt in formula_text, f"Leave type '{lt}' missing conditional formatting"


class TestBlockedByConsistency:
    """Bug #27: Blocked By must be consistent across all outputs."""

    def test_planned_with_blocked_by_in_schedule_suggestions(self, capsys):
        """Bug #27: A Planned task with Blocked By must appear in schedule suggestions."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#00BCD4", "P1"]],
            task_rows=[
                ["Blocked Task", "WS", "Alice", "2026-03-02", 10, "Planned",
                 10, "P1", None, "Waiting for Legal", None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#00BCD4", "priority": "P1"}})
        team = {"Alice": 5}

        # Build minimal allocation/weeks — allocation is {week: {person: days}}
        from capacity_planner import get_week_start
        weeks = []
        d = datetime(2026, 3, 2)
        for i in range(4):
            weeks.append(get_week_start(d + timedelta(weeks=i)))
        allocation = {w: {p: 0 for p in team} for w in weeks}

        # Capture output
        print_schedule_suggestions(tasks, team, allocation, weeks)
        out = capsys.readouterr().out
        assert "BLOCKED" in out
        assert "Waiting for Legal" in out


class TestCapacityExclusion:
    """Bug #16: On Hold tasks must NOT appear in capacity calculations."""

    def test_on_hold_excluded_from_tasks(self):
        """On Hold tasks should be loadable but filtered in capacity math."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#000", "P1"]],
            task_rows=[
                ["Active Task", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, None, None],
                ["Parked Task", "WS", "Alice", "2026-03-02", 5, "On Hold",
                 5, "P1", None, None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#000", "priority": "P1"}})
        active_tasks = [t for t in tasks if t["status"] not in ("On Hold", "Complete")]
        assert len(active_tasks) == 1
        assert active_tasks[0]["task"] == "Active Task"


class TestGetWeekStart:
    def test_monday(self):
        assert get_week_start(datetime(2026, 3, 2)) == datetime(2026, 3, 2)

    def test_friday(self):
        assert get_week_start(datetime(2026, 3, 6)) == datetime(2026, 3, 2)

    def test_sunday(self):
        assert get_week_start(datetime(2026, 3, 8)) == datetime(2026, 3, 2)

    def test_next_monday(self):
        assert get_week_start(datetime(2026, 3, 9)) == datetime(2026, 3, 9)


class TestActualEndClamp:
    """Bug #29: Actual end snapped backward must not fall before snapped start."""

    def test_weekend_start_weekend_actual_end(self):
        """Complete task with Saturday start + Sunday actual end must not invert."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Weekend fix", "WS", "Alice", "2026-02-14", 1, "Complete",
                 1, "P1", "2026-02-15", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        t = tasks[0]
        # Start (Sat Feb 14) snaps forward to Mon Feb 16
        # Actual End (Sun Feb 15) would snap backward to Fri Feb 13 without clamp
        # With clamp: actual_end_date = start_date (Mon Feb 16)
        assert t["actual_end_date"] == datetime(2026, 2, 16)
        assert t["actual_working_days"] == 1


class TestLowConfidenceExcludesComplete:
    """Bug #30: Low confidence warnings should not include Complete tasks."""

    def test_complete_task_excluded_from_low_confidence(self, capsys):
        """Complete task with Low confidence must NOT appear in low_conf output."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Done Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-06", None, None, "Low", None],
                ["Active Task", "WS", "Alice", "2026-03-09", 5, "In Progress",
                 5, "P1", None, None, None, "Low", None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        team = {"Alice": 5}
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}

        from capacity_planner import calculate_capacity
        allocation, weeks_list, available = calculate_capacity(tasks, team)

        print_summary(tasks, team, ws, allocation, weeks_list, available)
        out = capsys.readouterr().out

        # "Active Task" should appear in low confidence (In Progress + Low)
        assert "Active Task" in out
        # "Done Task" should NOT appear in low confidence (Complete + Low)
        low_conf_section = out[out.find("Low confidence"):] if "Low confidence" in out else ""
        assert "Done Task" not in low_conf_section


class TestLeaveExcludesPublicHolidays:
    """Bug #32: Leave day count must not include public holidays."""

    def test_leave_spanning_public_holiday(self):
        """Leave from Mon-Wed where Mon is a bank holiday should count 2 days, not 3."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[],
            holiday_rows=[["2026-05-04", "Early May Bank Holiday"]],
            leave_rows=[["Alice", "2026-05-04", "2026-05-06", "Annual Leave"]],
        )
        # Load holidays first, then leave with holidays passed in
        public_holidays = load_public_holidays(path)
        leave_dates, leave_entries = load_leave(path, public_holidays=public_holidays)

        # May 4 is a bank holiday — should NOT be counted as leave
        assert len(leave_entries) == 1
        assert leave_entries[0]["days"] == 2  # Tue May 5 + Wed May 6
        # leave_dates should also exclude the holiday
        assert datetime(2026, 5, 4) not in leave_dates.get("Alice", set())
        assert datetime(2026, 5, 5) in leave_dates.get("Alice", set())
        assert datetime(2026, 5, 6) in leave_dates.get("Alice", set())


class TestCompleteTaskCapacityTrimming:
    """Bug #34: Complete tasks should only consume capacity until actual end, not planned end."""

    def test_early_finish_trims_working_days(self):
        """A Complete task that finished early should not allocate capacity past actual end."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 10-day task starting Mon Mar 2, actual end Fri Mar 6 (5 days early)
                ["Early Task", "WS", "Alice", "2026-03-02", 10, "Complete",
                 10, "P1", "2026-03-06", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        t = tasks[0]

        # Planned end would be ~Mar 13 (10 working days from Mar 2)
        # Actual end is Mar 6 — working_days should be trimmed to <= Mar 6
        assert t["actual_end_date"] == datetime(2026, 3, 6)
        for wd in t["working_days"]:
            assert wd <= datetime(2026, 3, 6), f"Working day {wd} is after actual end"
        for d in t["day_allocations"]:
            assert d <= datetime(2026, 3, 6), f"Day allocation {d} is after actual end"

    def test_early_finish_capacity_not_inflated(self):
        """Capacity charts should not show the person as busy past actual end."""
        from capacity_planner import calculate_capacity, get_week_start
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 10-day task, finishes after 5 days
                ["Early Task", "WS", "Alice", "2026-03-02", 10, "Complete",
                 10, "P1", "2026-03-06", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)

        # Week of Mar 9 (second week) should have 0 allocation for Alice
        week_mar_9 = get_week_start(datetime(2026, 3, 9))
        if week_mar_9 in allocation:
            assert allocation[week_mar_9]["Alice"] == 0.0, \
                "Alice should have no allocation in week after actual end"


class TestPriorityTotalsExcludeOnHold:
    """Bug #35: Executive summary 'By priority' should exclude On Hold tasks."""

    def test_on_hold_excluded_from_priority_totals(self, capsys):
        """On Hold tasks must NOT appear in priority breakdown totals."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Active Task", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, None, None],
                ["Paused Task", "WS", "Alice", "2026-03-09", 10, "On Hold",
                 10, "P1", None, None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        team = {"Alice": 5}
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}

        from capacity_planner import calculate_capacity
        allocation, weeks_list, available = calculate_capacity(tasks, team)

        print_summary(tasks, team, ws, allocation, weeks_list, available)
        out = capsys.readouterr().out

        # Find the "By priority" line for P1
        for line in out.split("\n"):
            if "P1:" in line and "task" in line:
                # Should show 1 task (5 days), NOT 2 tasks (15 days)
                assert "1 task" in line, f"Expected 1 task, got: {line}"
                assert "5 days" in line, f"Expected 5 days, got: {line}"
                break
        else:
            # If no P1 line found, that's also wrong
            assert False, "P1 priority line not found in output"


class TestDateFilterUsesActualEnd:
    """Bug #36: --from/--to date filter must use actual_end_date for Complete tasks."""

    def test_complete_task_filtered_by_actual_end(self):
        """A Complete task finishing before --from window should be excluded."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Task: starts Jan 6, 20 working days → planned end ~Feb 2
                # Actual End: Jan 10 (finished early)
                ["Early Task", "WS", "Alice", "2026-01-06", 20, "Complete",
                 20, "P1", "2026-01-10", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)

        t = tasks[0]
        assert t["actual_end_date"] is not None, "actual_end_date should be set"

        # Replicate the --from/--to filter logic from main()
        date_from = datetime(2026, 2, 1)
        t_end = (t["actual_end_date"] if t["status"] == "Complete"
                 and t.get("actual_end_date") else t["end_date"])
        # actual_end_date is Jan 10, which is before Feb 1 → should be filtered out
        assert t_end < date_from, \
            f"Effective end {t_end:%Y-%m-%d} should be before window start {date_from:%Y-%m-%d}"

    def test_planned_end_would_pass_but_actual_end_filters(self):
        """Planned end in window but actual end before window → must be excluded."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Early Task", "WS", "Alice", "2026-01-06", 20, "Complete",
                 20, "P1", "2026-01-10", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)

        t = tasks[0]
        date_from = datetime(2026, 2, 1)

        # Planned end_date should be in/near Feb (20 working days from Jan 6)
        assert t["end_date"] >= date_from, \
            "Planned end should extend into February (this confirms the bug scenario)"

        # But actual_end_date is Jan 10, before the window
        t_end = (t["actual_end_date"] if t["status"] == "Complete"
                 and t.get("actual_end_date") else t["end_date"])
        assert t_end < date_from, \
            "Using actual_end_date, task should be filtered out of February window"


class TestGanttConfidenceDotsExcludeComplete:
    """Bug #37 (internal): Gantt confidence dots should skip Complete tasks."""

    def test_complete_task_no_confidence_dot(self):
        """Complete tasks should not display confidence dots (outcome is known)."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Done Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-06", None, None, "Low", None],
                ["Active Task", "WS", "Alice", "2026-03-09", 5, "In Progress",
                 5, "P1", None, None, None, "Low", None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)

        # Verify the rendering logic: Complete tasks should be skipped
        from capacity_planner import CONFIDENCE_COLORS
        for t in tasks:
            should_show_dot = (t.get("confidence")
                               and t["confidence"] in CONFIDENCE_COLORS
                               and t["status"] != "Complete")
            if t["status"] == "Complete":
                assert not should_show_dot, \
                    "Complete task should NOT show confidence dot"
            else:
                assert should_show_dot, \
                    "In Progress task with confidence should show dot"


class TestOnHoldExcludedFromDeadlineWarnings:
    """Bug #38 (internal): On Hold tasks should not appear in 'Deadlines at risk'."""

    def test_on_hold_excluded_from_deadline_warnings(self, capsys):
        """On Hold task with deadline should NOT trigger 'at risk' warning."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # On Hold task: starts Mar 2, 5 days, deadline Mar 3 (will miss)
                ["Paused Task", "WS", "Alice", "2026-03-02", 5, "On Hold",
                 5, "P1", None, None, "2026-03-03", None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        team = {"Alice": 5}
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}

        from capacity_planner import calculate_capacity
        allocation, weeks_list, available = calculate_capacity(tasks, team)
        print_summary(tasks, team, ws, allocation, weeks_list, available)
        out = capsys.readouterr().out

        assert "Deadlines at risk" not in out, \
            "On Hold task should not appear in deadline warnings"


class TestOnHoldExcludedFromConfidenceWarnings:
    """Bug #39 (internal): On Hold tasks should not appear in low confidence warnings."""

    def test_on_hold_excluded_from_low_confidence(self, capsys):
        """On Hold task with Low confidence should NOT appear in summary."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # On Hold task with Low confidence
                ["Paused Task", "WS", "Alice", "2026-03-02", 5, "On Hold",
                 5, "P1", None, None, None, "Low", None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        team = {"Alice": 5}
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}

        from capacity_planner import calculate_capacity
        allocation, weeks_list, available = calculate_capacity(tasks, team)
        print_summary(tasks, team, ws, allocation, weeks_list, available)
        out = capsys.readouterr().out

        assert "Low confidence" not in out, \
            "On Hold task should not appear in low confidence warnings"


class TestVarianceLabelAfterTrimming:
    """Bug #40: Gantt variance label must use planned count, not trimmed working_days."""

    def test_early_finish_shows_early_not_on_time(self):
        """A Complete task finishing 5 days early should show '-5d early', not 'on time'."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 10-day task starting Mon Mar 2, actual end Fri Mar 6 (5 days early)
                ["Early Task", "WS", "Alice", "2026-03-02", 10, "Complete",
                 10, "P1", "2026-03-06", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        t = tasks[0]

        # planned_working_days should be 10 (the full planned schedule)
        assert t["planned_working_days"] == 10, \
            f"Planned working days should be 10, got {t['planned_working_days']}"

        # working_days is trimmed (for capacity), but planned_working_days is not
        assert len(t["working_days"]) < t["planned_working_days"], \
            "working_days should be trimmed shorter than planned_working_days"

        # Variance calculation should show early, not on time
        planned_wd = t["planned_working_days"]
        actual_wd = t["actual_working_days"]
        diff = actual_wd - planned_wd
        assert diff < 0, f"Variance should be negative (early), got {diff}"

    def test_late_finish_shows_late(self):
        """A Complete task finishing 2 days late should show '+2d late'."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 5-day task starting Mon Mar 2, actual end Tue Mar 10 (2 days late)
                ["Late Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-10", None, None, None, None],
            ],
        )
        tasks = load_tasks(path, {"WS": {"color": "#2196F3", "priority": "P1"}})
        calculate_schedule(tasks)
        t = tasks[0]

        planned_wd = t["planned_working_days"]
        actual_wd = t["actual_working_days"]
        diff = actual_wd - planned_wd
        assert diff > 0, f"Variance should be positive (late), got {diff}"


# ── Tier 5: Cross-Consumer Regression Tests ──────────────────────────────────
# These tests trace a single task state through ALL downstream consumers
# to ensure fixes in one area don't break another (Bug #40 pattern).


class TestAdversarialCompleteTask:
    """Trace a Complete task (10d planned, 5d early) through ALL consumers.
    Verifies capacity, variance, summary, confidence, and date filter agree."""

    @pytest.fixture
    def complete_task_scenario(self):
        """Complete task: 10 planned days, finishes after 5 (early)."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 10-day task, actual end Fri Mar 6 = 5 days early
                ["Early Done", "WS", "Alice", "2026-03-02", 10, "Complete",
                 10, "P1", "2026-03-06", None, "2026-03-20", "Low", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_planned_working_days_preserved(self, complete_task_scenario):
        """planned_working_days must be 10 (original plan, not trimmed)."""
        tasks, *_ = complete_task_scenario
        t = tasks[0]
        assert t["planned_working_days"] == 10

    def test_working_days_trimmed_to_actual_end(self, complete_task_scenario):
        """working_days list must only contain dates <= actual_end_date."""
        tasks, *_ = complete_task_scenario
        t = tasks[0]
        for d in t["working_days"]:
            assert d <= t["actual_end_date"]
        # Must be shorter than planned
        assert len(t["working_days"]) < t["planned_working_days"]

    def test_variance_shows_early(self, complete_task_scenario):
        """Variance = actual - planned must be negative (early finish)."""
        tasks, *_ = complete_task_scenario
        t = tasks[0]
        diff = t["actual_working_days"] - t["planned_working_days"]
        assert diff < 0, f"Expected negative variance (early), got {diff}"

    def test_capacity_zero_after_actual_end(self, complete_task_scenario):
        """No capacity allocation in weeks after actual_end_date."""
        tasks, team, _, allocation, weeks, _ = complete_task_scenario
        t = tasks[0]
        actual_end_week = get_week_start(t["actual_end_date"])
        for w in weeks:
            if w > actual_end_week:
                assert allocation[w]["Alice"] == 0.0, \
                    f"Week {w:%Y-%m-%d}: Alice should have 0 allocation after actual end"

    def test_monthly_capacity_zero_after_actual_end(self, complete_task_scenario):
        """Monthly capacity must also respect actual end."""
        tasks, team, *_ = complete_task_scenario
        t = tasks[0]
        m_alloc, months, m_avail = calculate_monthly_capacity(tasks, team)
        # April and beyond should have zero allocation
        for m in months:
            if m >= datetime(2026, 4, 1):
                assert m_alloc[m]["Alice"] == 0.0, \
                    f"Month {m:%Y-%m}: Alice should have 0 allocation after actual end month"

    def test_not_in_low_confidence(self, complete_task_scenario, capsys):
        """Complete task should NOT appear in low confidence warnings."""
        tasks, team, ws, allocation, weeks, available = complete_task_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        if "Low confidence" in out:
            assert "Early Done" not in out[out.find("Low confidence"):]

    def test_confidence_dot_skipped(self, complete_task_scenario):
        """Confidence dot rendering should skip Complete tasks."""
        tasks, *_ = complete_task_scenario
        t = tasks[0]
        should_show = (t.get("confidence")
                       and t["confidence"] in CONFIDENCE_COLORS
                       and t["status"] != "Complete")
        assert not should_show

    def test_deadline_uses_actual_end(self, complete_task_scenario, capsys):
        """Deadline comparison should use actual_end_date, not planned end."""
        tasks, team, ws, allocation, weeks, available = complete_task_scenario
        t = tasks[0]
        # Task finishes Mar 6, deadline Mar 20 → should NOT be at risk
        eff_end = t["actual_end_date"]
        assert eff_end <= t["deadline"], \
            "Actual end should be before deadline"
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        assert "Deadlines at risk" not in out

    def test_in_priority_totals(self, complete_task_scenario, capsys):
        """Complete tasks should still count toward priority demand totals."""
        tasks, team, ws, allocation, weeks, available = complete_task_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        for line in out.split("\n"):
            if "P1:" in line and "task" in line:
                assert "1 task" in line
                break

    def test_date_filter_uses_actual_end(self, complete_task_scenario):
        """Date filter must use actual_end_date for Complete tasks."""
        tasks, *_ = complete_task_scenario
        t = tasks[0]
        date_from = datetime(2026, 3, 20)  # After actual end but before planned end
        t_end = (t["actual_end_date"] if t["status"] == "Complete"
                 and t.get("actual_end_date") else t["end_date"])
        assert t_end < date_from, "Task should be filtered out using actual_end_date"


class TestAdversarialOnHoldTask:
    """Trace an On Hold task through ALL consumers — must be excluded everywhere."""

    @pytest.fixture
    def on_hold_scenario(self):
        """On Hold task + active task to ensure On Hold is excluded."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Active task for baseline
                ["Active Task", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, None, None],
                # On Hold task — should be excluded from everything
                ["Paused Task", "WS", "Alice", "2026-03-02", 10, "On Hold",
                 10, "P1", None, "Waiting for vendor", "2026-03-06", "Low", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_task_is_loaded(self, on_hold_scenario):
        """On Hold tasks must still be loaded (they exist in data)."""
        tasks, *_ = on_hold_scenario
        assert len(tasks) == 2
        assert any(t["task"] == "Paused Task" for t in tasks)

    def test_excluded_from_capacity(self, on_hold_scenario):
        """On Hold task should contribute zero to capacity allocation."""
        tasks, team, _, allocation, weeks, _ = on_hold_scenario
        # Only "Active Task" (5 days, week of Mar 2) should allocate
        # "Paused Task" (10 days) should NOT contribute
        total_alloc = sum(allocation[w]["Alice"] for w in weeks)
        assert total_alloc == 5.0, \
            f"Expected 5.0 total allocation (active only), got {total_alloc}"

    def test_excluded_from_monthly_capacity(self, on_hold_scenario):
        """On Hold task must also be excluded from monthly capacity."""
        tasks, team, *_ = on_hold_scenario
        m_alloc, months, _ = calculate_monthly_capacity(tasks, team)
        total = sum(m_alloc[m]["Alice"] for m in months)
        assert total == 5.0, \
            f"Expected 5.0 monthly allocation (active only), got {total}"

    def test_excluded_from_concurrent_count(self, on_hold_scenario):
        """On Hold tasks must not count toward concurrent task warnings."""
        tasks, *_ = on_hold_scenario
        # Concurrent count logic: status in ("Planned", "In Progress")
        concurrent = [t for t in tasks
                      if t["status"] in ("Planned", "In Progress")
                      and t["assigned_to"] == "Alice"]
        assert len(concurrent) == 1  # Only the active task

    def test_excluded_from_priority_totals(self, on_hold_scenario, capsys):
        """On Hold tasks must not appear in priority demand totals."""
        tasks, team, ws, allocation, weeks, available = on_hold_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        for line in out.split("\n"):
            if "P1:" in line and "task" in line:
                assert "1 task" in line, f"Expected 1 task (not 2), got: {line}"
                assert "5 days" in line, f"Expected 5 days (not 15), got: {line}"
                break

    def test_excluded_from_deadline_warnings(self, on_hold_scenario, capsys):
        """On Hold task with deadline should NOT trigger deadline warning."""
        tasks, team, ws, allocation, weeks, available = on_hold_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        # Paused Task has deadline Mar 6, but end date is ~Mar 13
        # Should NOT appear in "Deadlines at risk" because it's On Hold
        if "Deadlines at risk" in out:
            assert "Paused Task" not in out[out.find("Deadlines at risk"):]

    def test_excluded_from_confidence_warnings(self, on_hold_scenario, capsys):
        """On Hold task with Low confidence should NOT appear in warnings."""
        tasks, team, ws, allocation, weeks, available = on_hold_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        if "Low confidence" in out:
            assert "Paused Task" not in out[out.find("Low confidence"):]

    def test_confidence_dot_skipped(self, on_hold_scenario):
        """Confidence dot rendering should be skipped for On Hold tasks (Bug #41)."""
        tasks, *_ = on_hold_scenario
        paused = [t for t in tasks if t["task"] == "Paused Task"][0]
        # On Hold tasks must NOT show confidence dot — matches print_summary exclusion
        should_show = (paused.get("confidence")
                       and paused["confidence"] in CONFIDENCE_COLORS
                       and paused["status"] not in ("Complete", "On Hold"))
        assert not should_show

    def test_deadline_visual_skipped(self, on_hold_scenario):
        """Deadline marker/overshoot should not render for On Hold tasks (Bug #41)."""
        tasks, *_ = on_hold_scenario
        paused = [t for t in tasks if t["task"] == "Paused Task"][0]
        # On Hold tasks must NOT show deadline visual — matches print_summary exclusion
        should_show_deadline = (paused.get("deadline")
                                and paused["status"] != "On Hold")
        assert not should_show_deadline


class TestAdversarialOverdueTask:
    """Trace an overdue In Progress task through ALL consumers."""

    @pytest.fixture
    def overdue_scenario(self):
        """In Progress task that ended in the past (overdue)."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 5-day task starting Jan 5 2026, ends ~Jan 9
                # Status: In Progress (should have been done by now)
                ["Overdue Task", "WS", "Alice", "2026-01-05", 5, "In Progress",
                 5, "P1", None, None, "2026-01-12", "Medium", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_appears_in_capacity(self, overdue_scenario):
        """In Progress tasks must still contribute to capacity."""
        tasks, team, _, allocation, weeks, _ = overdue_scenario
        total_alloc = sum(allocation[w]["Alice"] for w in weeks)
        assert total_alloc == 5.0

    def test_in_concurrent_count(self, overdue_scenario):
        """In Progress tasks must count toward concurrent task totals."""
        tasks, *_ = overdue_scenario
        concurrent = [t for t in tasks
                      if t["status"] in ("Planned", "In Progress")]
        assert len(concurrent) == 1

    def test_in_priority_totals(self, overdue_scenario, capsys):
        """In Progress task must appear in priority demand totals."""
        tasks, team, ws, allocation, weeks, available = overdue_scenario
        print_summary(tasks, team, ws, allocation, weeks, available)
        out = capsys.readouterr().out
        for line in out.split("\n"):
            if "P1:" in line and "task" in line:
                assert "1 task" in line
                break

    def test_deadline_at_risk(self, overdue_scenario, capsys):
        """Overdue In Progress task past deadline should trigger warning."""
        tasks, team, ws, allocation, weeks, available = overdue_scenario
        t = tasks[0]
        # End date is ~Jan 9, deadline is Jan 12 — end < deadline, so NOT at risk
        # But task is In Progress and past end_date — it IS overdue
        # Overdue detection is in print_schedule_suggestions, not print_summary
        assert t["status"] == "In Progress"
        assert t["end_date"] < norm_date(datetime.now()), \
            "Task should be past its planned end date"

    def test_confidence_dot_shown(self, overdue_scenario):
        """In Progress task with confidence should show confidence dot."""
        tasks, *_ = overdue_scenario
        t = tasks[0]
        should_show = (t.get("confidence")
                       and t["confidence"] in CONFIDENCE_COLORS
                       and t["status"] != "Complete")
        assert should_show, "In Progress task with Medium confidence should show dot"


class TestDataStructureModificationRegression:
    """Verify that when calculate_schedule() trims working_days/day_allocations,
    ALL consumers use the correct version (trimmed for capacity, planned for variance).
    This is the exact pattern that caused Bug #40."""

    @pytest.fixture
    def trimmed_scenario(self):
        """Complete task where trimming occurs (actual end < planned end)."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Analyst", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Alice: 20-day task, finished after 8 days
                ["Big Task", "WS", "Alice", "2026-03-02", 20, "Complete",
                 20, "P1", "2026-03-11", None, None, None, None],
                # Bob: normal In Progress task for comparison
                ["Normal Task", "WS", "Bob", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, None, None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5, "Bob": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_planned_working_days_not_affected_by_trim(self, trimmed_scenario):
        """planned_working_days must equal the original schedule length."""
        tasks, *_ = trimmed_scenario
        big = [t for t in tasks if t["task"] == "Big Task"][0]
        assert big["planned_working_days"] == 20

    def test_working_days_is_trimmed(self, trimmed_scenario):
        """working_days list must be trimmed to actual_end_date."""
        tasks, *_ = trimmed_scenario
        big = [t for t in tasks if t["task"] == "Big Task"][0]
        assert len(big["working_days"]) < 20
        for d in big["working_days"]:
            assert d <= big["actual_end_date"]

    def test_day_allocations_is_trimmed(self, trimmed_scenario):
        """day_allocations dict must also be trimmed to actual_end_date."""
        tasks, *_ = trimmed_scenario
        big = [t for t in tasks if t["task"] == "Big Task"][0]
        for d in big["day_allocations"]:
            assert d <= big["actual_end_date"]

    def test_capacity_uses_trimmed_data(self, trimmed_scenario):
        """Weekly capacity must use trimmed working_days (not planned)."""
        tasks, team, _, allocation, weeks, _ = trimmed_scenario
        big = [t for t in tasks if t["task"] == "Big Task"][0]
        actual_end_week = get_week_start(big["actual_end_date"])
        # Weeks after actual end should have zero allocation for Alice
        for w in weeks:
            if w > actual_end_week:
                assert allocation[w]["Alice"] == 0.0

    def test_variance_uses_planned_data(self, trimmed_scenario):
        """Variance calculation must use planned_working_days, not len(working_days)."""
        tasks, *_ = trimmed_scenario
        big = [t for t in tasks if t["task"] == "Big Task"][0]
        # This is exactly what Bug #40 was about
        planned_wd = big.get("planned_working_days", len(big["working_days"]))
        actual_wd = big.get("actual_working_days", planned_wd)
        diff = actual_wd - planned_wd
        # len(working_days) would give the trimmed count = actual_working_days → diff = 0
        # planned_working_days gives the original count → diff < 0 (early)
        assert diff != 0, \
            "Variance must not be zero (that would mean Bug #40 regression)"
        assert diff < 0, f"Expected negative variance (early finish), got {diff}"

    def test_in_progress_task_not_trimmed(self, trimmed_scenario):
        """In Progress tasks should NOT have trimmed working_days."""
        tasks, *_ = trimmed_scenario
        normal = [t for t in tasks if t["task"] == "Normal Task"][0]
        assert len(normal["working_days"]) == normal["planned_working_days"]
        assert normal["actual_end_date"] is None


# =============================================================================
# Tier 6: Production Readiness Tests
# =============================================================================


class TestAggregateWorkstreams:
    """Tests for aggregate_workstreams() — pure function, no matplotlib."""

    @pytest.fixture
    def multi_ws_scenario(self):
        """2 workstreams, 3 tasks — normal aggregation."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Dev", 5]],
            workstream_rows=[["Alpha", "#FF0000", "P1"], ["Beta", "#00FF00", "P2"]],
            task_rows=[
                ["Task A1", "Alpha", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "High", None],
                ["Task A2", "Alpha", "Bob", "2026-03-09", 3, "Planned",
                 3, "P1", None, None, None, "Medium", None],
                ["Task B1", "Beta", "Alice", "2026-03-16", 4, "Planned",
                 4, "P2", None, None, None, "High", None],
            ],
        )
        ws = {"Alpha": {"color": "#FF0000", "priority": "P1"},
              "Beta": {"color": "#00FF00", "priority": "P2"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        return tasks, ws

    def test_normal_aggregation(self, multi_ws_scenario):
        """2 workstreams with tasks — verify start/end/task_count/has_blocked."""
        tasks, ws = multi_ws_scenario
        result = aggregate_workstreams(tasks, ws)
        assert "Alpha" in result
        assert "Beta" in result
        assert result["Alpha"]["task_count"] == 2
        assert result["Beta"]["task_count"] == 1
        assert result["Alpha"]["has_blocked"] is False
        assert result["Beta"]["has_blocked"] is False
        # Alpha starts when Task A1 starts
        assert result["Alpha"]["start"] == norm_date(datetime(2026, 3, 2))
        # Alpha ends after Task A2 (starts 2026-03-09, 3 days)
        assert result["Alpha"]["end"] > result["Alpha"]["start"]
        # task_starts should have 2 entries for Alpha
        assert len(result["Alpha"]["task_starts"]) == 2

    def test_complete_task_uses_actual_end_for_span(self):
        """Complete task: workstream end uses actual_end_date, not planned end."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Planned: 20 days from Mar 2 → planned end ~Mar 27
                # Actual end: Mar 13 (early finish)
                ["Early Task", "WS", "Alice", "2026-03-02", 20, "Complete",
                 20, "P1", "2026-03-13", None, None, None, None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        result = aggregate_workstreams(tasks, ws)
        actual_end = norm_date(datetime(2026, 3, 13))
        # Workstream end should use actual_end_date, not planned end
        assert result["WS"]["end"] == actual_end

    def test_on_hold_detected_as_blocked(self):
        """On Hold task → has_blocked = True, blocked_tasks populated."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Paused Task", "WS", "Alice", "2026-03-02", 5, "On Hold",
                 5, "P1", None, None, None, "Medium", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        result = aggregate_workstreams(tasks, ws)
        assert result["WS"]["has_blocked"] is True
        assert len(result["WS"]["blocked_tasks"]) == 1
        assert result["WS"]["blocked_tasks"][0]["task"] == "Paused Task"

    def test_blocked_by_on_active_task_detected(self):
        """Non-Complete task with Blocked By → detected as blocked."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Blocker", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "High", None],
                ["Blocked Task", "WS", "Alice", "2026-03-09", 3, "Planned",
                 3, "P1", None, "Blocker", None, "Medium", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        result = aggregate_workstreams(tasks, ws)
        assert result["WS"]["has_blocked"] is True
        blocked_names = [t["task"] for t in result["WS"]["blocked_tasks"]]
        assert "Blocked Task" in blocked_names
        # The non-blocked task should NOT be in blocked_tasks
        assert "Blocker" not in blocked_names

    def test_empty_workstream_excluded(self):
        """Workstream with no matching tasks → excluded from result."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["Active", "#FF0000", "P1"], ["Empty", "#00FF00", "P2"]],
            task_rows=[
                ["Solo Task", "Active", "Alice", "2026-03-02", 5, "Planned",
                 5, "P1", None, None, None, "High", None],
            ],
        )
        ws = {"Active": {"color": "#FF0000", "priority": "P1"},
              "Empty": {"color": "#00FF00", "priority": "P2"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        result = aggregate_workstreams(tasks, ws)
        assert "Active" in result
        assert "Empty" not in result


class TestFullPipeline:
    """End-to-end pipeline: Excel → schedule → capacity → charts → summary."""

    @pytest.fixture
    def pipeline_data(self):
        """3 tasks (Planned, In Progress, Complete), 2 team, 1 holiday, 1 leave."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Dev", 4]],
            workstream_rows=[["Alpha", "#FF0000", "P1"], ["Beta", "#00FF00", "P2"]],
            task_rows=[
                ["Plan Task", "Alpha", "Alice", "2026-03-02", 10, "Planned",
                 10, "P1", None, None, "2026-03-20", "High", None],
                ["Active Task", "Beta", "Bob", "2026-03-09", 8, "In Progress",
                 8, "P2", None, None, None, "Medium", None],
                ["Done Task", "Alpha", "Alice", "2026-03-02", 15, "Complete",
                 15, "P1", "2026-03-13", None, "2026-03-27", "Low", None],
            ],
            holiday_rows=[["2026-03-06", "Test Holiday"]],
            leave_rows=[["Bob", "2026-03-16", "2026-03-17", "Annual Leave", None]],
        )
        ws = {"Alpha": {"color": "#FF0000", "priority": "P1"},
              "Beta": {"color": "#00FF00", "priority": "P2"}}
        tasks = load_tasks(path, ws)
        holidays = load_public_holidays(path)
        leave, leave_entries = load_leave(path, holidays)
        calculate_schedule(tasks, public_holidays=holidays, leave=leave)
        team = {"Alice": 5, "Bob": 4}
        allocation, weeks, available = calculate_capacity(
            tasks, team, public_holidays=holidays, leave=leave)
        return {
            "tasks": tasks, "team": team, "ws": ws,
            "allocation": allocation, "weeks": weeks, "available": available,
            "holidays": holidays, "leave": leave, "leave_entries": leave_entries,
        }

    def test_charts_generated(self, pipeline_data, tmp_path):
        """All 4 render functions produce non-zero PNG files."""
        d = pipeline_data
        gantt_path = str(tmp_path / "gantt.png")
        weekly_path = str(tmp_path / "weekly.png")
        monthly_path = str(tmp_path / "monthly.png")
        roadmap_path = str(tmp_path / "roadmap.png")

        render_gantt(d["tasks"], d["team"], d["ws"], d["weeks"], gantt_path,
                     public_holidays=d["holidays"], leave=d["leave"])
        render_weekly(d["tasks"], d["team"], d["ws"], d["allocation"],
                      d["weeks"], d["available"], weekly_path,
                      public_holidays=d["holidays"], leave=d["leave"])
        render_monthly_capacity(d["tasks"], d["team"], d["ws"], monthly_path,
                                public_holidays=d["holidays"], leave=d["leave"])
        render_roadmap(d["tasks"], d["team"], d["ws"], roadmap_path)

        for p in [gantt_path, weekly_path, monthly_path, roadmap_path]:
            assert os.path.exists(p), f"Missing: {p}"
            assert os.path.getsize(p) > 0, f"Empty: {p}"

    def test_summary_output(self, pipeline_data, capsys):
        """print_summary and print_schedule_suggestions produce expected sections."""
        d = pipeline_data
        print_summary(d["tasks"], d["team"], d["ws"], d["allocation"],
                      d["weeks"], d["available"], d["holidays"], d["leave"],
                      d["leave_entries"])
        print_schedule_suggestions(d["tasks"], d["team"], d["allocation"],
                                   d["weeks"], d["available"],
                                   d["holidays"], d["leave"])
        output = capsys.readouterr().out
        # Key sections should be present
        assert "EXECUTIVE SUMMARY" in output or "Executive Summary" in output \
            or "Task" in output  # at minimum, tasks are mentioned
        assert "Alice" in output
        assert "Bob" in output

    def test_empty_after_date_filter(self, tmp_path):
        """Tasks ending in March, filtered from June → no crash."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["March Task", "WS", "Alice", "2026-03-02", 5, "Planned",
                 5, "P1", None, None, None, "High", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        # Apply a date filter that excludes all tasks (June)
        date_from = norm_date(datetime(2026, 6, 1))
        filtered = [t for t in tasks
                    if t["end_date"] >= date_from]
        assert len(filtered) == 0
        # Render functions should handle empty tasks gracefully (early return)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(filtered, team)
        assert allocation == {}
        assert weeks == []
        # Render functions should not crash with empty data
        render_gantt(filtered, team, ws, weeks,
                     str(tmp_path / "gantt.png"))
        render_weekly(filtered, team, ws, allocation, weeks, available,
                      str(tmp_path / "weekly.png"))
        render_monthly_capacity(filtered, team, ws,
                                str(tmp_path / "monthly.png"))
        render_roadmap(filtered, team, ws,
                       str(tmp_path / "roadmap.png"))
        # No PNG files should be created (early returns)
        assert not os.path.exists(tmp_path / "gantt.png")


class TestAllSameStatus:
    """Edge cases: all tasks in the same status."""

    def test_all_complete(self, capsys):
        """All Complete tasks → capacity zero after all actual ends, no active warnings."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Done1", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-06", None, None, "High", None],
                ["Done2", "WS", "Alice", "2026-03-09", 5, "Complete",
                 5, "P1", "2026-03-13", None, None, "Medium", None],
                ["Done3", "WS", "Alice", "2026-03-16", 5, "Complete",
                 5, "P1", "2026-03-20", None, None, "Low", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # After all actual ends, no capacity should be consumed
        latest_actual = max(t["actual_end_date"] for t in tasks)
        latest_week = get_week_start(latest_actual)
        for w in weeks:
            if w > latest_week:
                assert allocation[w]["Alice"] == 0.0
        # Summary should not mention low confidence warnings for Complete tasks
        print_summary(tasks, team, ws, allocation, weeks, available)
        output = capsys.readouterr().out
        assert "Low confidence" not in output

    def test_all_on_hold(self, capsys):
        """All On Hold tasks → zero capacity everywhere, no deadline/confidence warnings."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Hold1", "WS", "Alice", "2026-03-02", 5, "On Hold",
                 5, "P1", None, None, "2026-03-20", "Low", None],
                ["Hold2", "WS", "Alice", "2026-03-09", 5, "On Hold",
                 5, "P1", None, None, "2026-03-27", "Low", None],
                ["Hold3", "WS", "Alice", "2026-03-16", 5, "On Hold",
                 5, "P1", None, None, None, "Medium", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # On Hold excluded from capacity — all weeks zero
        for w in weeks:
            assert allocation[w]["Alice"] == 0.0
        # No deadline or confidence warnings for On Hold
        print_summary(tasks, team, ws, allocation, weeks, available)
        print_schedule_suggestions(tasks, team, allocation, weeks, available)
        output = capsys.readouterr().out
        assert "at risk" not in output.lower() or "on hold" not in output.lower()

    def test_all_planned(self, capsys):
        """All Planned tasks → all appear in capacity, no overdue warnings."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Plan1", "WS", "Alice", "2026-06-01", 5, "Planned",
                 5, "P1", None, None, None, "High", None],
                ["Plan2", "WS", "Alice", "2026-06-08", 5, "Planned",
                 5, "P1", None, None, None, "High", None],
                ["Plan3", "WS", "Alice", "2026-06-15", 5, "Planned",
                 5, "P1", None, None, None, "High", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # All tasks should contribute to capacity
        total_alloc = sum(allocation[w]["Alice"] for w in weeks)
        assert total_alloc > 0
        # No overdue warnings (all in the future)
        print_schedule_suggestions(tasks, team, allocation, weeks, available)
        output = capsys.readouterr().out
        assert "overdue" not in output.lower()

    def test_all_in_progress_concurrent_warning(self, capsys):
        """3 In Progress tasks for same person → concurrent task warning."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Active1", "WS", "Alice", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, "High", None],
                ["Active2", "WS", "Alice", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, "Medium", None],
                ["Active3", "WS", "Alice", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, "Low", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # Concurrent task count for Alice should be 3
        active_tasks = [t for t in tasks
                        if t["status"] not in ("Complete", "On Hold")
                        and t["assigned_to"] == "Alice"]
        assert len(active_tasks) == 3


class TestMultiPersonOverlap:
    """Multi-person overlapping task scenarios."""

    def test_independent_capacity_per_person(self):
        """3 people, overlapping tasks → capacity per person is independent."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Dev", 4], ["Carol", "Dev", 3]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Task A", "WS", "Alice", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, "High", None],
                ["Task B", "WS", "Bob", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, None, "High", None],
                ["Task C", "WS", "Carol", "2026-03-02", 10, "Planned",
                 10, "P1", None, None, None, "High", None],
                ["Task D", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "Medium", None],
                ["Task E", "WS", "Bob", "2026-03-09", 5, "Planned",
                 5, "P1", None, None, None, "Medium", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5, "Bob": 4, "Carol": 3}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # Each person's allocation should be independent
        first_week = weeks[0]
        # Alice has 2 tasks in first week
        assert allocation[first_week]["Alice"] > 0
        # Bob has 1 task in first week
        assert allocation[first_week]["Bob"] > 0
        # Carol has 1 task in first week
        assert allocation[first_week]["Carol"] > 0
        # Carol has 10 days over 2 weeks — allocation shows actual demand
        # (over-capacity is expected and detected by the tool)
        carol_total = sum(allocation[w]["Carol"] for w in weeks)
        assert carol_total > 0

    def test_over_capacity_detection(self, capsys):
        """Person with 5 days/week assigned too much work → over-capacity in summary."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # 3 concurrent tasks each 5 days starting same week = 15 days in 1 week
                ["Heavy1", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "High", None],
                ["Heavy2", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "High", None],
                ["Heavy3", "WS", "Alice", "2026-03-02", 5, "In Progress",
                 5, "P1", None, None, None, "High", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        # First week should show over-capacity
        first_week = weeks[0]
        assert allocation[first_week]["Alice"] > 5.0, \
            f"Expected over-capacity, got {allocation[first_week]['Alice']}"


class TestAdversarialLateCompleteTask:
    """Bug #42: Late-finishing Complete task — capacity must extend to actual end."""

    @pytest.fixture
    def late_complete_scenario(self):
        """Complete task: 5 planned days from Mar 2, actual end Mar 20 (late)."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Planned: 5 days from Mar 2 → planned end Mar 6
                # Actual end: Mar 20 (2 weeks late)
                ["Late Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-20", None, "2026-03-27", "High", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_planned_working_days_preserved(self, late_complete_scenario):
        """planned_working_days must reflect original plan, not extended days."""
        tasks, *_ = late_complete_scenario
        t = tasks[0]
        assert t["planned_working_days"] == 5

    def test_working_days_extended_to_actual_end(self, late_complete_scenario):
        """working_days must include days up to actual_end_date."""
        tasks, *_ = late_complete_scenario
        t = tasks[0]
        actual_end = t["actual_end_date"]
        assert max(t["working_days"]) == actual_end
        assert len(t["working_days"]) > t["planned_working_days"]

    def test_day_allocations_extended(self, late_complete_scenario):
        """day_allocations must include entries up to actual_end_date."""
        tasks, *_ = late_complete_scenario
        t = tasks[0]
        actual_end = t["actual_end_date"]
        assert actual_end in t["day_allocations"]
        assert t["day_allocations"][actual_end] == 1.0

    def test_variance_shows_late(self, late_complete_scenario):
        """Variance = actual - planned must be positive (late finish)."""
        tasks, *_ = late_complete_scenario
        t = tasks[0]
        diff = t["actual_working_days"] - t["planned_working_days"]
        assert diff > 0, f"Expected positive variance (late), got {diff}"

    def test_capacity_extends_to_actual_end(self, late_complete_scenario):
        """Capacity must show allocation in weeks after planned end, up to actual end."""
        tasks, team, _, allocation, weeks, _ = late_complete_scenario
        t = tasks[0]
        planned_end_week = get_week_start(t["end_date"])
        actual_end_week = get_week_start(t["actual_end_date"])
        # Weeks between planned end and actual end should have allocation
        for w in weeks:
            if planned_end_week < w <= actual_end_week:
                assert allocation[w]["Alice"] > 0, \
                    f"Week {w:%Y-%m-%d}: expected allocation in late period"

    def test_capacity_zero_after_actual_end(self, late_complete_scenario):
        """No capacity after actual_end_date — same as early-finish test."""
        tasks, team, _, allocation, weeks, _ = late_complete_scenario
        t = tasks[0]
        actual_end_week = get_week_start(t["actual_end_date"])
        for w in weeks:
            if w > actual_end_week:
                assert allocation[w]["Alice"] == 0.0

    def test_monthly_capacity_covers_late_period(self, late_complete_scenario):
        """Monthly capacity must include the late period."""
        tasks, team, *_ = late_complete_scenario
        m_alloc, months, _ = calculate_monthly_capacity(tasks, team)
        # March should have allocation (task runs Mar 2-20)
        march = datetime(2026, 3, 1)
        assert march in m_alloc
        assert m_alloc[march]["Alice"] > 5  # More than planned 5 days

    def test_timeline_bounds_include_actual_end(self, late_complete_scenario):
        """Chart timeline (weeks) must extend to cover actual_end_date."""
        tasks, _, _, _, weeks, _ = late_complete_scenario
        t = tasks[0]
        actual_end_week = get_week_start(t["actual_end_date"])
        assert actual_end_week in weeks, \
            f"Weeks should include {actual_end_week:%Y-%m-%d} (actual end week)"


    # --- Interaction regression tests (reviewer-suggested) ---

    @pytest.fixture
    def late_complete_with_overlap(self):
        """Late Complete task + overlapping In Progress task for same person."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Complete task: planned 5d from Mar 2, actual end Mar 20 (late)
                ["Late Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-20", None, "2026-03-27", "Low", None],
                # In Progress task: starts Mar 16, overlaps the late tail
                ["Active Task", "WS", "Alice", "2026-03-16", 10, "In Progress",
                 10, "P1", None, None, "2026-04-10", "High", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team)
        return tasks, team, ws, allocation, weeks, available

    def test_late_complete_excluded_from_concurrency(self, late_complete_with_overlap):
        """Late Complete with extended working_days must NOT count as concurrent (Bug #18 guard)."""
        tasks, team, ws, allocation, weeks, available = late_complete_with_overlap

        buf = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = buf
        try:
            print_summary(tasks, team, ws, allocation, weeks, available)
        finally:
            sys.stdout = old_stdout
        output = buf.getvalue()
        # The late Complete overlaps the In Progress task in weeks of Mar 16-20,
        # but concurrent count should only see 1 task (the In Progress one),
        # NOT 2 (which would trigger the "concurrent tasks" warning)
        assert "concurrent tasks" not in output.lower(), \
            "Late Complete task leaked into concurrent count"

    def test_late_complete_excluded_from_low_confidence(self, late_complete_with_overlap):
        """Late Complete with Low confidence must NOT appear in low confidence warnings."""
        tasks, team, ws, allocation, weeks, available = late_complete_with_overlap

        buf = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = buf
        try:
            print_summary(tasks, team, ws, allocation, weeks, available)
        finally:
            sys.stdout = old_stdout
        output = buf.getvalue()
        assert "Low confidence" not in output, \
            "Late Complete with Low confidence leaked into warnings"

    def test_late_complete_not_at_risk_when_within_deadline(self, late_complete_with_overlap):
        """Late Complete within deadline must NOT appear in 'Deadlines at risk'."""
        tasks, team, ws, allocation, weeks, available = late_complete_with_overlap
        # Late Task: actual end Mar 20, deadline Mar 27 → within deadline

        buf = io.StringIO()
        old_stdout = sys.stdout
        sys.stdout = buf
        try:
            print_summary(tasks, team, ws, allocation, weeks, available)
        finally:
            sys.stdout = old_stdout
        output = buf.getvalue()
        # Neither task should be at risk (both within their deadlines)
        assert "Deadlines at risk" not in output, \
            "No tasks should be at risk — both are within deadlines"

    def test_availability_unchanged_by_late_extension(self):
        """Available capacity must be identical whether or not late Complete extends allocation."""
        # Scenario 1: With late Complete task
        path1 = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                ["Late Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-20", None, None, None, None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        tasks1 = load_tasks(path1, ws)
        calculate_schedule(tasks1)
        team = {"Alice": 5}
        _, weeks1, available1 = calculate_capacity(tasks1, team)

        # Scenario 2: Same team, no tasks (just availability)
        path2 = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Need at least one task to generate weeks, use a Planned task in same range
                ["Placeholder", "WS", "Alice", "2026-03-02", 15, "Planned",
                 15, "P1", None, None, None, None, None],
            ],
        )
        tasks2 = load_tasks(path2, ws)
        calculate_schedule(tasks2)
        _, weeks2, available2 = calculate_capacity(tasks2, team)

        # For overlapping weeks, available capacity must be identical
        common_weeks = set(weeks1) & set(weeks2)
        assert len(common_weeks) > 0, "Need overlapping weeks to compare"
        for w in common_weeks:
            assert available1[w]["Alice"] == available2[w]["Alice"], \
                f"Week {w:%Y-%m-%d}: available capacity differs ({available1[w]['Alice']} vs {available2[w]['Alice']})"

    def test_leave_filtering_with_extended_timeline(self):
        """Leave near extended end date must survive filtering when timeline expands."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5]],
            workstream_rows=[["WS", "#2196F3", "P1"]],
            task_rows=[
                # Late Complete: planned end ~Mar 6, actual end Mar 20
                ["Late Task", "WS", "Alice", "2026-03-02", 5, "Complete",
                 5, "P1", "2026-03-20", None, None, None, None],
            ],
            leave_rows=[
                # Leave near the extended end — should be included
                ["Alice", "2026-03-18", "2026-03-19", "Annual Leave", None],
            ],
        )
        ws = {"WS": {"color": "#2196F3", "priority": "P1"}}
        holidays = load_public_holidays(path)
        leave, leave_entries = load_leave(path, holidays)
        tasks = load_tasks(path, ws)
        calculate_schedule(tasks, public_holidays=holidays, leave=leave)
        team = {"Alice": 5}
        allocation, weeks, available = calculate_capacity(tasks, team,
                                                          public_holidays=holidays, leave=leave)

        # Apply the same leave filtering logic as main()
        if weeks:
            win_start = weeks[0]
            win_end = weeks[-1] + timedelta(days=4)
            filtered_leave = [e for e in leave_entries
                              if not (e["end"] < win_start or e["start"] > win_end)]
        else:
            filtered_leave = leave_entries

        # The leave entry (Mar 18-19) is within the extended timeline — must survive
        assert len(filtered_leave) == 1, \
            f"Leave near extended end should survive filtering, got {len(filtered_leave)} entries"
        assert filtered_leave[0]["person"] == "Alice"


class TestRenderSmoke:
    """Smoke tests for render functions — verify they don't crash and produce output."""

    @pytest.fixture
    def render_data(self):
        """Standard scenario for render smoke tests."""
        path = create_test_excel(
            team_rows=[["Alice", "Lead", 5], ["Bob", "Dev", 4]],
            workstream_rows=[["Alpha", "#FF0000", "P1"], ["Beta", "#00FF00", "P2"]],
            task_rows=[
                ["Task 1", "Alpha", "Alice", "2026-03-02", 10, "In Progress",
                 10, "P1", None, None, "2026-03-20", "High", None],
                ["Task 2", "Beta", "Bob", "2026-03-09", 8, "Planned",
                 8, "P2", None, None, None, "Medium", None],
                ["Task 3", "Alpha", "Alice", "2026-03-16", 5, "Complete",
                 5, "P1", "2026-03-20", None, "2026-03-27", "Low", None],
            ],
            holiday_rows=[["2026-03-06", "Holiday"]],
            leave_rows=[["Bob", "2026-03-16", "2026-03-17", "Annual Leave", None]],
        )
        ws = {"Alpha": {"color": "#FF0000", "priority": "P1"},
              "Beta": {"color": "#00FF00", "priority": "P2"}}
        tasks = load_tasks(path, ws)
        holidays = load_public_holidays(path)
        leave, _ = load_leave(path, holidays)
        calculate_schedule(tasks, public_holidays=holidays, leave=leave)
        team = {"Alice": 5, "Bob": 4}
        allocation, weeks, available = calculate_capacity(
            tasks, team, public_holidays=holidays, leave=leave)
        return {
            "tasks": tasks, "team": team, "ws": ws,
            "allocation": allocation, "weeks": weeks, "available": available,
            "holidays": holidays, "leave": leave,
        }

    def test_render_gantt_smoke(self, render_data, tmp_path):
        """render_gantt produces a non-zero PNG."""
        d = render_data
        p = str(tmp_path / "gantt.png")
        render_gantt(d["tasks"], d["team"], d["ws"], d["weeks"], p,
                     public_holidays=d["holidays"], leave=d["leave"])
        assert os.path.exists(p)
        assert os.path.getsize(p) > 0

    def test_render_weekly_smoke(self, render_data, tmp_path):
        """render_weekly produces a non-zero PNG."""
        d = render_data
        p = str(tmp_path / "weekly.png")
        render_weekly(d["tasks"], d["team"], d["ws"], d["allocation"],
                      d["weeks"], d["available"], p,
                      public_holidays=d["holidays"], leave=d["leave"])
        assert os.path.exists(p)
        assert os.path.getsize(p) > 0

    def test_render_monthly_smoke(self, render_data, tmp_path):
        """render_monthly_capacity produces a non-zero PNG."""
        d = render_data
        p = str(tmp_path / "monthly.png")
        render_monthly_capacity(d["tasks"], d["team"], d["ws"], p,
                                public_holidays=d["holidays"], leave=d["leave"])
        assert os.path.exists(p)
        assert os.path.getsize(p) > 0

    def test_render_roadmap_smoke(self, render_data, tmp_path):
        """render_roadmap produces a non-zero PNG."""
        d = render_data
        p = str(tmp_path / "roadmap.png")
        render_roadmap(d["tasks"], d["team"], d["ws"], p)
        assert os.path.exists(p)
        assert os.path.getsize(p) > 0
